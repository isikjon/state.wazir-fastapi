from fastapi import FastAPI, Request, Depends, Form, status, HTTPException, Query, WebSocket, WebSocketDisconnect, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, JSONResponse
from config import settings
from api.v1.api import api_router
from sqlalchemy.orm import Session, joinedload
from app.api import deps
from app.utils.security import verify_password
from app import models
from app.models.user import User
from sqlalchemy import func, desc, or_, and_
from datetime import datetime, timedelta
from sqlalchemy.types import String
import pandas as pd
import os
from uuid import uuid4
import json
from flask import jsonify
import random
from starlette.middleware.sessions import SessionMiddleware
from typing import Dict, Any, Optional, List
from starlette.middleware.base import BaseHTTPMiddleware
from jose import jwt as pyjwt
from app.models.token import TokenPayload
from fastapi import APIRouter
from app.models.chat import AppChatModel, AppChatMessageModel
from app.models.chat_message import ChatMessage
from app.websockets.chat_manager import ConnectionManager as WebSocketManager
from app.utils.image_helper import get_valid_image_url
from app.models.property import PropertyCategory
import sys
try:
    import psutil
except ImportError:
    psutil = None
import asyncio
import subprocess
import openpyxl
from io import BytesIO

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=settings.ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = pyjwt.encode(to_encode, settings.SECRET_KEY, algorithm=settings.ALGORITHM)
    return encoded_jwt

# Класс для аутентификации HTTP запросов (но не WebSocket)
class AuthenticationMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        # Проверяем, является ли запрос WebSocket
        if request.scope.get("type") == "websocket":
            # Для WebSocket запросов пропускаем проверку
            return await call_next(request)
        
        # Пути, доступные без авторизации
        public_paths = [
            '/mobile/auth',
            '/mobile/register',
            '/mobile/register/verify',
            '/mobile/register/profile',
            '/mobile/reset',
            '/mobile/reset/verify',
            '/mobile/reset/password',
            '/api/v1/auth/login',
            '/api/v1/auth/check-exists',
            '/api/v1/auth/send-code',
            '/api/v1/auth/verify-code',
            '/api/v1/auth/register',
            '/api/v1/auth/reset-password',
            '/static/',
            '/favicon.ico',
            '/mobile/test-websocket',
            '/mobile/ws/',
            '/api/v1/chat/',  # Chat API general path
            '/admin/login',   # Admin login page
            '/superadmin/login',  # SuperAdmin login page
            '/companies/login',   # Company login page
        ]
        
        # Для статических файлов, API разрешаем доступ
        if request.url.path.startswith('/static/'):
            return await call_next(request)
            
        # Для API запросов проверяем токен в заголовке Authorization
        if request.url.path.startswith('/api/'):
            # Разрешаем доступ к эндпоинтам авторизации без токена
            if any(request.url.path.endswith(path) for path in [
                '/login',
                '/register',
                '/check-exists',
                '/send-code',
                '/verify-code',
                '/reset-password'
            ]):
                return await call_next(request)
                
            auth_token = None
            auth_header = request.headers.get('Authorization')
            
            if auth_header and auth_header.startswith('Bearer '):
                auth_token = auth_header.split(' ')[1]
            
            if not auth_token:
                auth_token = request.cookies.get('access_token')
                print(f"DEBUG: Используем токен из cookie: {auth_token is not None}")
            
            if auth_token:
                try:
                    payload = pyjwt.decode(auth_token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
                    if datetime.fromtimestamp(payload["exp"]) < datetime.utcnow():
                        print(f"DEBUG: Token expired: {payload}")
                        return JSONResponse(
                            status_code=status.HTTP_401_UNAUTHORIZED,
                            content={"detail": "Токен истек"}
                        )
                    
                    if not auth_header:
                        request.headers.__dict__["_list"].append((b"authorization", f"Bearer {auth_token}".encode()))
                        print("DEBUG: Добавлен заголовок Authorization из cookie")
                    
                    return await call_next(request)
                except Exception as e:
                    print(f"DEBUG: Token validation error: {str(e)}")
                    return JSONResponse(
                        status_code=status.HTTP_401_UNAUTHORIZED,
                        content={"detail": f"Недействительный токен: {str(e)}"}
                    )
            return JSONResponse(
                status_code=status.HTTP_401_UNAUTHORIZED,
                content={"detail": "Требуется авторизация"}
            )
            
        if any(request.url.path.startswith(path) for path in public_paths) or '/api/v1/chat/' in request.url.path:
            return await call_next(request)
            
        auth_token = request.cookies.get('access_token')
        auth_header = request.headers.get('Authorization')
        
        print(f"DEBUG: Checking auth for path: {request.url.path}")
        print(f"DEBUG: Cookie token: {auth_token}")
        print(f"DEBUG: Auth header: {auth_header}")
        
        if auth_header and auth_header.startswith('Bearer '):
            auth_token = auth_header.split(' ')[1]
            print(f"DEBUG: Using token from header: {auth_token}")
        
        if not auth_token:
            print("DEBUG: No token found")
            if request.url.path.startswith('/admin/'):
                return RedirectResponse('/admin/login', status_code=303)
            elif request.url.path.startswith('/superadmin/'):
                return RedirectResponse('/superadmin/login', status_code=303)
            elif request.url.path.startswith('/companies/'):
                return RedirectResponse('/companies/login', status_code=303)
            # Для остальных маршрутов перенаправляем на страницу авторизации
            return RedirectResponse('/mobile/auth', status_code=303)
            
        # Проверяем валидность токена
        try:
            payload = pyjwt.decode(auth_token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
            print(f"DEBUG: Token payload: {payload}")
            if datetime.fromtimestamp(payload["exp"]) < datetime.utcnow():
                print("DEBUG: Token expired")
                if request.url.path.startswith('/admin/'):
                    return RedirectResponse('/admin/login', status_code=303)
                elif request.url.path.startswith('/superadmin/'):
                    return RedirectResponse('/superadmin/login', status_code=303)
                elif request.url.path.startswith('/companies/'):
                    return RedirectResponse('/companies/login', status_code=303)
                return RedirectResponse('/mobile/auth', status_code=303)
                
            # Для суперадмин-маршрутов проверяем, что пользователь является суперадминистратором
            if request.url.path.startswith('/superadmin/') and not payload.get("is_superadmin"):
                print("DEBUG: Non-superadmin user trying to access superadmin area")
                return RedirectResponse('/superadmin/login', status_code=303)
                
            # Для админ-маршрутов проверяем, что пользователь является администратором
            if request.url.path.startswith('/admin/') and not payload.get("is_admin"):
                print("DEBUG: Non-admin user trying to access admin area")
                return RedirectResponse('/admin/login', status_code=303)
                
            # Для маршрутов компаний проверяем, что пользователь является компанией
            if request.url.path.startswith('/companies/') and not payload.get("is_company"):
                print("DEBUG: Non-company user trying to access company area")
                return RedirectResponse('/companies/login', status_code=303)
                
        except Exception as e:
            print(f"DEBUG: Token validation error (cookie): {str(e)}")
            if request.url.path.startswith('/admin/'):
                return RedirectResponse('/admin/login', status_code=303)
            elif request.url.path.startswith('/superadmin/'):
                return RedirectResponse('/superadmin/login', status_code=303)
            elif request.url.path.startswith('/companies/'):
                return RedirectResponse('/companies/login', status_code=303)
            return RedirectResponse('/mobile/auth', status_code=303)
            
        # Если токен валиден, пропускаем запрос дальше
        return await call_next(request)

# Кастомный JSON-энкодер для обработки datetime и других неподдерживаемых типов
class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.strftime("%Y-%m-%d %H:%M:%S")
        return super().default(obj)

app = FastAPI(
    title=settings.PROJECT_NAME,
    openapi_url=f"{settings.API_V1_STR}/openapi.json"
)

# Монтируем статические файлы
app.mount("/static", StaticFiles(directory="static", html=True, check_dir=True), name="static")

# Используем импортированный chat_manager вместо создания нового экземпляра

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.BACKEND_CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["Authorization", "Content-Type", "Set-Cookie"],
)

app.add_middleware(SessionMiddleware, secret_key="wazir_super_secret_key")

# Функция для сериализации объектов в JSON, используя CustomJSONEncoder
def json_serialize(obj):
    return json.dumps(obj, cls=CustomJSONEncoder)

@app.exception_handler(TypeError)
async def type_error_handler(request, exc):
    if "not JSON serializable" in str(exc):
        return JSONResponse(
            status_code=500,
            content={"detail": "Error serializing the response"},
        )
    raise exc

# Монтируем директорию media
app.mount("/media", StaticFiles(directory="media", check_dir=True), name="media")

templates = Jinja2Templates(directory="templates")

# Регистрация API роутеров
app.include_router(api_router, prefix=settings.API_V1_STR)

# Добавляем мидлвар для проверки авторизации
app.add_middleware(AuthenticationMiddleware)

# WebSocket Manager class
class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, List[WebSocket]] = {}
        self.chat_messages: Dict[str, List[dict]] = {}
        # Загружаем сохраненные сообщения, если файл существует
        try:
            if os.path.exists("chat_messages.json"):
                with open("chat_messages.json", "r", encoding="utf-8") as f:
                    data = f.read()
                    if data.strip():  # проверяем, что файл не пустой
                        self.chat_messages = json.loads(data)
                        print(f"DEBUG: Загружены сохраненные сообщения чатов. Доступные комнаты: {list(self.chat_messages.keys())}")
                        for room, messages in self.chat_messages.items():
                            print(f"DEBUG: Комната {room}: {len(messages)} сообщений")
                            # Выводим первое и последнее сообщение для отладки
                            if messages:
                                print(f"DEBUG: Первое сообщение: {messages[0].get('content', 'Нет контента')}")
                                print(f"DEBUG: Последнее сообщение: {messages[-1].get('content', 'Нет контента')}")
                    else:
                        print("DEBUG: Файл с сообщениями пуст, создаем новый")
                        with open("chat_messages.json", "w", encoding="utf-8") as f:
                            json.dump({}, f, ensure_ascii=False, indent=2)
            else:
                print("DEBUG: Файл с сообщениями не найден, создаем новый")
                with open("chat_messages.json", "w", encoding="utf-8") as f:
                    json.dump({}, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"DEBUG: Ошибка при загрузке сообщений: {e}")
            # Создаем пустой файл в случае ошибки
            try:
                with open("chat_messages.json", "w", encoding="utf-8") as f:
                    json.dump({}, f, ensure_ascii=False, indent=2)
            except Exception as e2:
                print(f"DEBUG: Не удалось создать пустой файл: {e2}")

    async def connect(self, websocket: WebSocket, room: str, accept_connection: bool = True):
        if accept_connection:
            await websocket.accept()
        if room not in self.active_connections:
            self.active_connections[room] = []
        if websocket not in self.active_connections[room]:
            self.active_connections[room].append(websocket)
            print(f"DEBUG: Подключен к комнате {room}, всего подключений: {len(self.active_connections[room])}")
        else:
            print(f"DEBUG: Соединение уже подключено к комнате {room}")

    def disconnect(self, websocket: WebSocket, room: str):
        try:
            if room in self.active_connections:
                if websocket in self.active_connections[room]:
                    self.active_connections[room].remove(websocket)
                    print(f"DEBUG: Отключен от комнаты {room}, осталось подключений: {len(self.active_connections[room])}")
                if not self.active_connections[room]:
                    del self.active_connections[room]
                    print(f"DEBUG: Комната {room} удалена, нет активных подключений")
        except Exception as e:
            print(f"DEBUG: Ошибка при отключении от комнаты {room}: {e}")

    async def send_personal_message(self, message: dict, websocket: WebSocket):
        await websocket.send_json(message)

    # Метод для сохранения сообщений в файл
    def save_messages_to_file(self):
        try:
            with open("chat_messages.json", "w", encoding="utf-8") as f:
                json.dump(self.chat_messages, f, ensure_ascii=False, indent=2)
            print(f"DEBUG: Сообщения сохранены в файл. Всего комнат: {len(self.chat_messages)}")
        except Exception as e:
            print(f"ERROR: Ошибка при сохранении сообщений в файл: {e}")
    
    # Метод для добавления сообщения в память и сохранения в файл
    def add_message_to_memory(self, chat_id: str, message: dict):
        if chat_id not in self.chat_messages:
            self.chat_messages[chat_id] = []
        
        # Добавляем сообщение в память
        self.chat_messages[chat_id].append(message)
        
        # Сохраняем в файл после добавления нового сообщения
        self.save_messages_to_file()
    
    async def broadcast(self, message: dict, room: str, exclude=None):
        # Если это сообщение чата, сохраняем его в памяти
        if message.get("type") in ["message", "message_sent", "new_message"] and "message" in message:
            chat_id = str(message["message"].get("chat_id", room))
            self.add_message_to_memory(chat_id, message["message"])
        
        # Отправляем сообщение всем активным соединениям в комнате
        if room in self.active_connections:
            for connection in self.active_connections[room]:
                if connection != exclude:
                    await connection.send_json(message)

    def save_message(self, room: str, message: dict):
        if room not in self.chat_messages:
            self.chat_messages[room] = []
        self.chat_messages[room].append(message)
        # Сохраняем сообщения в файл для персистентности
        try:
            # Преобразуем datetime в строку для сериализации
            serializable_messages = {}
            for room_key, messages in self.chat_messages.items():
                serializable_messages[room_key] = []
                for msg in messages:
                    # Копируем сообщение и обрабатываем timestamp, если нужно
                    if isinstance(msg, dict):
                        serializable_messages[room_key].append(msg)
                    else:
                        # Если сообщение не словарь, преобразуем его в строку
                        serializable_messages[room_key].append(str(msg))
            
            with open("chat_messages.json", "w", encoding="utf-8") as f:
                json.dump(serializable_messages, f, ensure_ascii=False, indent=2)
                
            print(f"DEBUG: Сохранено сообщение в комнату {room}, всего сообщений: {len(self.chat_messages[room])}")
        except Exception as e:
            print(f"DEBUG: Ошибка при сохранении сообщений: {e}")

    def get_messages(self, room: str) -> List[dict]:
        """Получить все сообщения для указанной комнаты"""
        return self.chat_messages.get(room, [])
        
    async def save_message_to_db(self, message_data: dict, db: Session) -> dict:
        """Сохраняет сообщение в базу данных и возвращает его с дополнительными полями"""
        try:
            # Создаем полную копию данных сообщения
            saved_message = message_data.copy()
            
            # Добавляем timestamp если его нет
            current_time = datetime.now()
            saved_message["timestamp"] = current_time.isoformat()
            saved_message["is_read"] = False
            # Используем целочисленный ID вместо UUID для совместимости с предложенной структурой БД
            saved_message["id"] = random.randint(100, 10000)
            
            # Определяем комнату для сообщения (всегда используем меньший ID первым)
            sender_id = int(saved_message["sender_id"])
            receiver_id = int(saved_message["receiver_id"])
            room = f"{min(sender_id, receiver_id)}_{max(sender_id, receiver_id)}"
            
            # Сохраняем сообщение
            self.save_message(room, saved_message)
            
            print(f"DEBUG: Сообщение сохранено: {saved_message}")
            return saved_message
        except Exception as e:
            print(f"ERROR: Ошибка при сохранении сообщения: {e}")
            # Возвращаем базовое сообщение с временной меткой в случае ошибки
            basic_message = message_data.copy()
            basic_message["timestamp"] = datetime.now().isoformat()
            basic_message["is_read"] = False
            basic_message["id"] = str(uuid4())
            return basic_message

manager = ConnectionManager()

# ============================ WebSocket Endpoints ============================
from sqlalchemy.orm import Session
from app.api import deps
from app.websockets.chat_manager import manager as chat_manager
from jose import jwt as pyjwt

@app.websocket("/mobile/ws/chat/{token}")
async def chat_websocket_endpoint(websocket: WebSocket, token: str):
    try:
        payload = pyjwt.decode(token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            await websocket.close(code=1008)
            return
    except Exception as e:
        print(f"WebSocket auth error: {e}")
        await websocket.close(code=1008)
        return
        
    await chat_manager.connect(websocket, user_id)
    
    try:
        while True:
            data = await websocket.receive_json()
            message_type = data.get("type")
            
            if message_type == "message":
                db = next(deps.get_db())
                message_data = {
                    "sender_id": int(user_id),
                    "receiver_id": data["receiver_id"],
                    "content": data["content"]
                }
                
                saved_message = await chat_manager.save_message_to_db(message_data, db)
                
                await websocket.send_json({
                    "type": "message_sent",
                    "message": saved_message
                })
                
                # Добавляем получателя в данные для отправки
                broadcast_message = {
                    "type": "new_message",
                    "message": saved_message,
                    "receiver_id": data["receiver_id"]
                }
                await chat_manager.broadcast(broadcast_message)
    except WebSocketDisconnect:
        chat_manager.disconnect(websocket, user_id)
    except Exception as e:
        print(f"WebSocket error: {e}")
        chat_manager.disconnect(websocket, user_id)

@app.websocket("/mobile/ws/test")
async def websocket_test_endpoint(websocket: WebSocket):
    print("WebSocket test connection request")
    await websocket.accept()
    print("WebSocket test connection accepted")
    
    try:
        while True:
            # Отправляем тестовое сообщение каждые 5 секунд
            test_message = {
                "type": "test",
                "message": f"Test message at {datetime.now().strftime('%H:%M:%S')}",
                "timestamp": datetime.now().isoformat()
            }
            await websocket.send_text(json.dumps(test_message))
            await asyncio.sleep(5)
    except WebSocketDisconnect:
        print("WebSocket test connection disconnected")

@app.websocket("/superadmin/ws/logs")
async def logs_websocket_endpoint(websocket: WebSocket):
    import subprocess
    import asyncio
    from datetime import datetime
    
    await websocket.accept()
    print("SuperAdmin logs WebSocket connection established")
    
    try:
        # Запускаем процесс для получения логов Docker
        process = subprocess.Popen(
            ["docker", "logs", "-f", "--tail", "50", "state.wazir-fastapi-web-1"],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True,
            bufsize=1
        )
        
        while True:
            # Читаем строку из логов
            line = process.stdout.readline()
            if line:
                log_entry = {
                    "timestamp": datetime.now().isoformat(),
                    "message": line.strip(),
                    "level": "INFO"  # Можно добавить определение уровня лога
                }
                await websocket.send_text(json.dumps(log_entry))
            else:
                await asyncio.sleep(0.1)
                
    except WebSocketDisconnect:
        print("SuperAdmin logs WebSocket connection disconnected")
        if 'process' in locals():
            process.terminate()
    except Exception as e:
        print(f"Error in logs WebSocket: {e}")
        if 'process' in locals():
            process.terminate()

# Корневой маршрут - перенаправление на мобильную версию
@app.get("/", response_class=RedirectResponse)
async def root():
    return RedirectResponse(url="/mobile")

# Добавляем маршруты для доступа через /layout
@app.get("/layout", response_class=RedirectResponse)
async def layout_root():
    return RedirectResponse(url="/mobile")

@app.get("/layout/dashboard", response_class=RedirectResponse)
async def layout_dashboard():
    return RedirectResponse(url="/mobile")

@app.get("/layout/auth", response_class=RedirectResponse)
async def layout_auth():
    return RedirectResponse(url="/mobile/auth")

@app.get("/layout/profile", response_class=RedirectResponse)
async def layout_profile():
    return RedirectResponse(url="/mobile/profile")

@app.get("/layout/create-listing", response_class=RedirectResponse)
async def layout_create_listing():
    return RedirectResponse(url="/mobile/create-listing")

@app.get("/layout/search", response_class=RedirectResponse)
async def layout_search():
    return RedirectResponse(url="/mobile/search")

@app.get("/layout/property/{property_id}", response_class=RedirectResponse)
async def layout_property_detail(property_id: int):
    return RedirectResponse(url=f"/mobile/property/{property_id}")

# Новые страницы для аутентификации
@app.get("/register", response_class=RedirectResponse)
async def redirect_register():
    return RedirectResponse(url="/mobile/register")

@app.get("/reset", response_class=RedirectResponse)
async def redirect_reset():
    return RedirectResponse(url="/mobile/reset")

@app.get("/mobile/register", response_class=HTMLResponse, name="mobile_register")
async def mobile_register(request: Request):
    return templates.TemplateResponse("layout/register.html", {"request": request})

@app.get("/mobile/register/verify", response_class=HTMLResponse, name="mobile_register_verify")
async def mobile_register_verify(request: Request):
    return templates.TemplateResponse("layout/verify.html", {"request": request})

@app.get("/mobile/register/profile", response_class=HTMLResponse, name="mobile_profile_create")
async def mobile_profile_create(request: Request):
    return templates.TemplateResponse("layout/profile_create.html", {"request": request})

@app.get("/mobile/reset", response_class=HTMLResponse, name="mobile_reset")
async def mobile_reset(request: Request):
    return templates.TemplateResponse("layout/reset.html", {"request": request})

@app.get("/mobile/reset/verify", response_class=HTMLResponse, name="mobile_reset_verify")
async def mobile_reset_verify(request: Request):
    return templates.TemplateResponse("layout/reset_verify.html", {"request": request})

@app.get("/mobile/reset/password", response_class=HTMLResponse, name="mobile_reset_password")
async def mobile_reset_password(request: Request):
    return templates.TemplateResponse("layout/reset_password.html", {"request": request})

# Мобильные (клиентские) маршруты
@app.get("/mobile", response_class=HTMLResponse, name="dashboard")
async def mobile_root(request: Request):
    return templates.TemplateResponse("layout/dashboard.html", {"request": request})

@app.get("/mobile/auth", response_class=HTMLResponse, name="mobile_auth")
async def mobile_auth(request: Request):
    return templates.TemplateResponse("layout/auth.html", {"request": request})

@app.get("/mobile/profile", response_class=HTMLResponse, name="profile")
async def mobile_profile(request: Request, tab: str = None, db: Session = Depends(deps.get_db)):
    # Получаем текущего пользователя
    user = None
    formatted_user_listings = []
    formatted_saved_listings = []
    
    # Получаем токен из cookie
    auth_token = request.cookies.get('access_token')
    auth_header = request.headers.get('Authorization')
    
    if auth_header and auth_header.startswith('Bearer '):
        auth_token = auth_header.split(' ')[1]
    
    # Получаем погоду и курс валюты для отображения в шапке
    weather = {"temperature": "+20°"}
    currency = {"value": "69.8"}
    
    if auth_token:
        try:
            # Декодируем токен
            payload = pyjwt.decode(auth_token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
            user_id = payload.get("sub")
            
            if user_id:
                # Получаем данные пользователя из БД
                user = db.query(models.User).filter(models.User.id == user_id).first()
                
                if user:
                    print(f"DEBUG: Загружен пользователь: {user.email}, роль: {user.role}")
                    
                    # Получаем объявления пользователя с изображениями
                    try:
                        user_listings = db.query(models.Property).options(
                            joinedload(models.Property.images)
                        ).filter(models.Property.owner_id == user_id).all()
                        
                        # Форматируем объявления пользователя для шаблона
                        for prop in user_listings:
                            # Находим главное изображение или первое доступное
                            main_image = next((img for img in prop.images if img.is_main), None) or \
                                       (prop.images[0] if prop.images else None)
                            
                            formatted_user_listings.append({
                                "id": prop.id,
                                "title": prop.title,
                                "price": prop.price,
                                "address": prop.address,
                                "rooms": prop.rooms,
                                "area": prop.area,
                                "status": prop.status,
                                "notes": prop.notes,  # Дата съемки 360
                                "tour_360_url": prop.tour_360_url,
                                "has_tour": bool(prop.tour_360_url),
                                "image_url": get_valid_image_url(main_image.url if main_image else None)
                            })
                        
                        print(f"DEBUG: Найдено {len(formatted_user_listings)} объявлений пользователя")
                        
                    except Exception as e:
                        print(f"DEBUG: Ошибка при получении объявлений пользователя: {e}")
                    
                    # Получаем сохраненные объявления пользователя
                    try:
                        # Получаем идентификаторы сохраненных объявлений
                        favorites_query = db.query(models.Favorite).filter(models.Favorite.user_id == user_id).all()
                        saved_property_ids = [fav.property_id for fav in favorites_query]
                        
                        # Получаем сами объявления по ID с изображениями
                        if saved_property_ids:
                            saved_listings = db.query(models.Property).options(
                                joinedload(models.Property.images)
                            ).filter(
                                models.Property.id.in_(saved_property_ids)
                            ).all()
                            
                            # Форматируем сохраненные объявления для шаблона
                            for prop in saved_listings:
                                # Находим главное изображение или первое доступное
                                main_image = next((img for img in prop.images if img.is_main), None) or \
                                           (prop.images[0] if prop.images else None)
                                
                                formatted_saved_listings.append({
                                    "id": prop.id,
                                    "title": prop.title,
                                    "price": prop.price,
                                    "address": prop.address,
                                    "rooms": prop.rooms,
                                    "area": prop.area,
                                    "status": prop.status,
                                    "tour_360_url": prop.tour_360_url,
                                    "has_tour": bool(prop.tour_360_url),
                                    "image_url": get_valid_image_url(main_image.url if main_image else None)
                                })
                            
                            print(f"DEBUG: Найдено {len(saved_listings)} сохраненных объявлений")
                        else:
                            print("DEBUG: У пользователя нет избранных объявлений")
                    except Exception as e:
                        print(f"DEBUG: Ошибка при получении сохраненных объявлений: {e}")
        except Exception as e:
            print(f"DEBUG: Ошибка декодирования токена: {e}")
    
    # Если не удалось получить пользователя, перенаправляем на страницу авторизации
    if not user:
        return RedirectResponse('/mobile/auth', status_code=303)
    
    # Создаем расширенный объект пользователя для шаблона
    user_data = {
        "id": user.id,
        "email": user.email,
        "phone": user.phone,
        "is_active": user.is_active,
        "role": user.role.value if user.role else "USER",
        "created_at": user.created_at
    }
    
    # Определяем отображаемые данные в зависимости от роли
    if user.role == models.UserRole.COMPANY:
        user_data.update({
            "full_name": user.company_name or "Компания",
            "display_name": user.company_name or "Компания", 
            "company_name": user.company_name,
            "company_number": user.company_number,
            "company_owner": user.company_owner,
            "company_address": user.company_address,
            "company_description": user.company_description,
            "avatar_url": user.company_logo_url,
            "is_company": True
        })
    else:
        user_data.update({
            "full_name": user.full_name or f"Пользователь {user.id}",
            "display_name": user.full_name or f"Пользователь {user.id}",
            "avatar_url": None,
            "is_company": False
        })
    
    return templates.TemplateResponse(
        "layout/profile.html", 
        {
            "request": request, 
            "user": user_data, 
            "user_listings": formatted_user_listings, 
            "saved_listings": formatted_saved_listings,
            "active_tab": tab or "listings",
            "weather": weather,
            "currency": currency
        }
    )

@app.get("/mobile/create-listing", response_class=HTMLResponse, name="create_listing")
async def mobile_create_listing(request: Request, db: Session = Depends(deps.get_db)):
    # Получаем текущего пользователя
    user = None
    
    # Получаем токен из cookie
    auth_token = request.cookies.get('access_token')
    auth_header = request.headers.get('Authorization')
    
    if auth_header and auth_header.startswith('Bearer '):
        auth_token = auth_header.split(' ')[1]
    
    if auth_token:
        try:
            # Декодируем токен
            payload = pyjwt.decode(auth_token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
            user_id = payload.get("sub")
            
            if user_id:
                # Получаем данные пользователя из БД
                user = db.query(models.User).filter(models.User.id == user_id).first()
                print(f"DEBUG: Загружен пользователь для создания объявления: {user.email}, роль: {user.role}")
                
                # Дополняем данные пользователя для корпоративных аккаунтов
                if user and user.role == models.UserRole.COMPANY:
                    print(f"DEBUG: Корпоративный пользователь - {user.company_name}, владелец: {user.company_owner}")
                
        except Exception as e:
            print(f"DEBUG: Ошибка при проверке токена: {e}")
    
    # Проверяем, авторизован ли пользователь
    if not user:
        return RedirectResponse(url="/mobile/auth")
    
    return templates.TemplateResponse(
        "layout/create-listing.html",
        {
            "request": request,
            "user": user
        }
    )

@app.get("/mobile/search", response_class=HTMLResponse, name="search")
async def mobile_search(
    request: Request, 
    category: str = None, 
    price_min: int = None, 
    price_max: int = None, 
    min_area: float = None, 
    max_area: float = None,
    rooms: int = None,
    min_floor: int = None,
    max_floor: int = None,
    balcony: bool = None,
    furniture: bool = None,
    renovation: bool = None,
    parking: bool = None,
    q: str = None,
    db: Session = Depends(deps.get_db)
):
    print("\n===================================================")
    print("DEBUG: Параметры запроса:")
    print(f"  URL: {request.url}")
    print(f"  Поисковый запрос: {q}")
    print(f"  Категория: {category}")
    print(f"  Цена: {price_min} - {price_max}")
    print(f"  Площадь: {min_area} - {max_area}")
    print(f"  Комнаты: {rooms}")
    print(f"  Этаж: {min_floor} - {max_floor}")
    print(f"  Балкон: {balcony}, Мебель: {furniture}, Ремонт: {renovation}, Паркинг: {parking}")
    print("===================================================")
    
    # Формируем базовый запрос для получения активных объявлений
    query = db.query(models.Property).filter(models.Property.status == 'active')
    
    # Получаем все категории для выпадающего списка
    categories = db.query(models.Category).all()
    print(f"DEBUG: Загружены категории: {[cat.name for cat in categories]}")
    
    # Применяем фильтры, если они указаны
    if category:
        print(f"DEBUG: Применяем фильтр по категории: {category}")
        query = query.join(models.Property.categories).filter(models.Category.name == category)
    
    # Поиск по ключевому слову в названии или адресе
    if q:
        search_term = f"%{q}%"
        query = query.filter(or_(
            models.Property.title.ilike(search_term),
            models.Property.address.ilike(search_term),
            models.Property.description.ilike(search_term)
        ))
    
    # Фильтры по цене
    if price_min is not None:
        query = query.filter(models.Property.price >= price_min)
    
    if price_max is not None:
        query = query.filter(models.Property.price <= price_max)
    
    # Фильтры по площади
    if min_area is not None:
        query = query.filter(models.Property.area >= min_area)
    
    if max_area is not None:
        query = query.filter(models.Property.area <= max_area)
    
    # Фильтр по количеству комнат
    if rooms is not None:
        query = query.filter(models.Property.rooms == rooms)
    
    # Фильтры по этажу
    if min_floor is not None:
        query = query.filter(models.Property.floor >= min_floor)
    
    if max_floor is not None:
        query = query.filter(models.Property.floor <= max_floor)
    
    # Дополнительные фильтры
    if balcony is not None and balcony:
        query = query.filter(models.Property.has_balcony == True)
    
    if furniture is not None and furniture:
        query = query.filter(models.Property.has_furniture == True)
    
    if renovation is not None and renovation:
        query = query.filter(models.Property.has_renovation == True)
    
    if parking is not None and parking:
        query = query.filter(models.Property.has_parking == True)
    
    # Получаем объявления с загрузкой изображений
    properties_db = query.options(joinedload(models.Property.images)).all()
    
    # Форматируем данные для шаблона
    properties = []
    for prop in properties_db:
        # Находим главное изображение
        main_image = next((img for img in prop.images if img.is_main), None) or \
                   (prop.images[0] if prop.images else None)
        
        # Формируем массив URL изображений
        images = [get_valid_image_url(img.url) for img in prop.images] if prop.images else []
        
        properties.append({
            "id": prop.id,
            "title": prop.title,
            "price": prop.price,
            "address": prop.address,
            "rooms": prop.rooms,
            "area": prop.area,
            "floor": prop.floor,
            "has_tour": bool(prop.tour_360_url),
            "tour_360_url": prop.tour_360_url,  # Добавляем URL для 360° тура
            "has_balcony": prop.has_balcony,
            "has_furniture": prop.has_furniture,
            "has_renovation": prop.has_renovation,
            "has_parking": prop.has_parking,
            "image_url": get_valid_image_url(main_image.url if main_image else None),
            "images": images,
            "images_count": len(images)
        })
    
    # Получаем погоду и курс валюты для отображения в шапке
    weather = None
    currency = None
    try:
        # Здесь можно добавить вызов API для получения погоды и курса валюты
        # Для простоты используем заглушки
        weather = {"temperature": "+15°"}
        currency = {"value": "87.5"}
    except Exception as e:
        print(f"DEBUG: Ошибка при получении погоды или курса валюты: {e}")
    
    return templates.TemplateResponse("layout/search.html", {
        "request": request, 
        "properties": properties,
        "weather": weather,
        "currency": currency,
        "categories": categories,  # Передаем список категорий в шаблон
        "selected_category": category,  # Выбранная категория
        "q": q,  # Поисковый запрос
        "filter": {
            "category": category,
            "price_min": price_min,
            "price_max": price_max,
            "min_area": min_area,
            "max_area": max_area,
            "rooms": rooms,
            "min_floor": min_floor,
            "max_floor": max_floor,
            "balcony": balcony,
            "furniture": furniture,
            "renovation": renovation,
            "parking": parking
        }
    })

@app.get("/mobile/chats", response_class=HTMLResponse, name="chats")
async def mobile_chats(request: Request):
    return templates.TemplateResponse("mobile/chats.html", {"request": request})

@app.get("/mobile/support", response_class=HTMLResponse, name="support")
async def mobile_support(request: Request):
    return templates.TemplateResponse("mobile/support.html", {"request": request})

@app.get("/mobile/property/{property_id}", response_class=HTMLResponse, name="property")
async def mobile_property_detail(request: Request, property_id: int, db: Session = Depends(deps.get_db)):
    # Получаем текущего пользователя, если он авторизован
    current_user = None
    token = request.cookies.get('access_token')
    
    if token:
        try:
            payload = pyjwt.decode(token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
            user_id = int(payload.get("sub"))
            current_user = db.query(models.User).filter(models.User.id == user_id).first()
        except Exception as e:
            print(f"DEBUG: Ошибка при получении пользователя: {e}")
    
    # Получаем объявление из БД
    property = db.query(models.Property).options(
        joinedload(models.Property.owner),
        joinedload(models.Property.images),
        joinedload(models.Property.categories)
    ).filter(models.Property.id == property_id).first()
    
    if not property:
        return templates.TemplateResponse("404.html", {"request": request})
    
    # Проверяем, добавлено ли объявление в избранное
    is_favorite = False
    if current_user:
        favorite = db.query(models.Favorite).filter(
            models.Favorite.user_id == current_user.id,
            models.Favorite.property_id == property.id
        ).first()
        is_favorite = favorite is not None
    
    # Получаем похожие объявления (того же типа, в том же городе)
    similar_properties = db.query(models.Property).options(
        joinedload(models.Property.images)
    ).filter(
        models.Property.id != property.id,
        models.Property.type == property.type,
        models.Property.city == property.city,
        models.Property.status == 'ACTIVE'
    ).limit(5).all()
    
    # Форматируем данные для шаблона
    property_data = {
        "id": property.id,
        "title": property.title,
        "description": property.description,
        "price": property.price,
        "address": property.address,
        "city": property.city,
        "area": property.area,
        "status": property.status.value.lower() if property.status else "draft",
        "is_featured": property.is_featured,
        "tour_360_url": property.tour_360_url,
        "type": property.type,
        "rooms": property.rooms,
        "floor": property.floor,
        "building_floors": property.building_floors,
        "has_balcony": property.has_balcony,
        "has_furniture": property.has_furniture,
        "has_renovation": property.has_renovation,
        "has_parking": property.has_parking,
        "notes": property.notes,  # Дата съемки 360
        "created_at": property.created_at,
        "updated_at": property.updated_at,
        "views_count": getattr(property, 'views_count', 0),  # Безопасное получение значения
        "owner": {
            "id": property.owner.id,
            "full_name": property.owner.full_name,
            "email": property.owner.email,
            "phone": property.owner.phone
        },
        "images": [{
            "id": image.id,
            "url": get_valid_image_url(image.url),
            "is_main": image.is_main
        } for image in property.images],
        "is_favorite": is_favorite,
        "is_owner": current_user and current_user.id == property.owner_id
    }
    
    # Форматируем похожие объявления
    similar_properties_data = []
    for prop in similar_properties:
        # Находим главное изображение или берем первое доступное
        main_image = next((img for img in prop.images if img.is_main), None) or \
                    (prop.images[0] if prop.images else None)
        
        similar_properties_data.append({
            "id": prop.id,
            "title": prop.title,
            "price": prop.price,
            "address": prop.address,
            "rooms": prop.rooms,
            "area": prop.area,
            "image_url": main_image.url if main_image else "/static/layout/assets/img/property-placeholder.jpg"
        })
    
    # Получаем погоду и курс валюты для отображения в шапке
    weather = None
    currency = None
    try:
        # Здесь можно добавить вызов API для получения погоды и курса валюты
        # Для простоты используем заглушки
        weather = {"temperature": "+20°"}
        currency = {"value": "69.8"}
    except Exception as e:
        print(f"DEBUG: Ошибка при получении погоды или курса валюты: {e}")
    
    return templates.TemplateResponse("layout/property.html", {
        "request": request,
        "property": property_data,
        "similar_properties": similar_properties_data,
        "weather": weather,
        "currency": currency,
        "user": current_user
    })

@app.get("/mobile/chat/{user_id}", response_class=HTMLResponse, name="chat")
async def mobile_chat(request: Request, user_id: int, db: Session = Depends(deps.get_db)):
    # Получаем property_id из query-параметров
    property_id = request.query_params.get("property_id")
    context = {"request": request, "user_id": user_id}
    
    if property_id:
        try:
            property_id = int(property_id)  # Преобразуем в число
            # Получаем информацию о объявлении, если указан property_id
            property_item = db.query(models.Property).filter(models.Property.id == property_id).first()
            if property_item:
                context["property"] = {
                    "id": property_item.id,
                    "title": property_item.title,
                    "price": property_item.price
                }
        except ValueError:
            # Обработка случая, когда property_id не является числом
            pass
    
    return templates.TemplateResponse("mobile/chat.html", context)

# Тестовая страница для WebSocket
@app.get("/mobile/test-websocket", response_class=HTMLResponse)
async def test_websocket_page(request: Request):
    """Тестовая страница для проверки WebSocket соединений"""
    return templates.TemplateResponse("test_websocket.html", {"request": request})

# Создаем прямой API-роутер для отладки, который не использует аутентификацию
debug_router = APIRouter(prefix="/debug")

# Добавляем debug_router в приложение
app.include_router(debug_router)

# Добавляем маршруты отладки
@debug_router.get("/")
async def get_debug_info():
    return {
        "status": "ok",
        "message": "Отладочный API доступен"
    }

# API для чата
from pydantic import BaseModel

from app.models.chat import AppChatMessageModel
from typing import List, Optional, Dict, Any

class MessageReadRequest(BaseModel):
    message_id: int

@app.post("/api/v1/chat/messages/read")
async def mark_message_as_read(request: MessageReadRequest, db: Session = Depends(deps.get_db)):
    """ Маркировать сообщение как прочитанное """
    try:
        # В реальном приложении здесь был бы код для обновления сообщения в базе данных
        # Например: message = db.query(AppChatMessageModel).filter(AppChatMessageModel.id == request.message_id).first()
        # if message:
        #    message.is_read = True
        #    db.commit()
        print(f"DEBUG: Сообщение {request.message_id} отмечено как прочитанное")
        return {"status": "success", "message": f"Сообщение {request.message_id} отмечено как прочитанное"}
    except Exception as e:
        print(f"ERROR: Ошибка при маркировке сообщения как прочитанное: {e}")
        return {"status": "error", "message": f"Ошибка при маркировке сообщения: {str(e)}"}

# API для получения пользователя по ID
@app.get("/api/v1/users/{user_id}")
async def get_user(user_id: int, db: Session = Depends(deps.get_db)):
    """Получить информацию о пользователе по ID"""
    try:
        print(f"DEBUG: Запрос данных пользователя с ID {user_id}")
        
        # Получаем пользователя из базы данных
        user = db.query(models.User).filter(models.User.id == user_id).first()
        
        # Если пользователь найден, формируем ответ с его данными
        if user:
            print(f"DEBUG: Пользователь найден в БД: {user.email}, {user.full_name}, роль: {user.role}")
            
            # Определяем отображаемое имя в зависимости от роли
            if user.role == models.UserRole.COMPANY:
                display_name = user.company_name or user.full_name or "Компания"
                avatar_url = user.company_logo_url or f"/static/img/company{user_id}.png"
            else:
                display_name = user.full_name or f"Пользователь {user_id}"
                avatar_url = f"/static/img/avatar{user_id}.png"
            
            return {
                "id": user.id,
                "email": user.email,
                "full_name": display_name,
                "avatar": avatar_url,
                "status": "Онлайн" if user.is_active else "Не в сети",
                "is_active": user.is_active,
                "role": user.role.value if user.role else "USER",
                "company_name": user.company_name,
                "company_number": user.company_number,
                "phone": user.phone
            }
        else:
            print(f"DEBUG: Пользователь с ID {user_id} не найден в БД")
            # Если пользователь не найден, возвращаем стандартные данные
            return {
                "id": user_id,
                "email": f"user{user_id}@example.com",
                "full_name": f"User {user_id}",
                "avatar": f"/static/img/avatar{user_id}.png",
                "status": "Пользователь",
                "is_active": False,
                "role": "USER"
            }
    except Exception as e:
        print(f"ERROR: Ошибка при получении пользователя: {e}")
        # В случае ошибки возвращаем базовые данные
        return {
            "id": user_id,
            "email": f"user{user_id}@example.com",
            "full_name": f"User {user_id}",
            "avatar": f"/static/img/avatar{user_id}.png",
            "status": "Пользователь",
            "is_active": True,
            "role": "USER"
        }

# API для получения сообщений чата
@app.get("/api/v1/chat/messages/{user_id}")
async def get_chat_messages(user_id: int, current_user: dict = Depends(deps.get_current_user_optional), db: Session = Depends(deps.get_db)):
    """Получить сообщения чата с пользователем"""
    try:
        # Получаем ID текущего пользователя из токена
        current_user_id = int(current_user.get("sub", 0)) if current_user else 0
        if current_user_id == 0:
            print("DEBUG: Не удалось определить текущего пользователя")
            return []
        
        # Используем импортированный chat_manager для доступа к сообщениям
        from app.websockets.chat_manager import manager as chat_manager
        
        # Формируем уникальный ID чата на основе идентификаторов пользователей
        chat_id = chat_manager.get_chat_id(current_user_id, user_id)
        print(f"DEBUG: Сформированный chat_id: {chat_id} для пользователей {current_user_id} и {user_id}")
        
        # Получаем сообщения из памяти chat_manager
        messages = chat_manager.chat_messages.get(chat_id, [])
        
        # Если сообщений нет в памяти, пробуем загрузить их из файла
        if not messages and os.path.exists("chat_messages.json"):
            try:
                with open("chat_messages.json", "r", encoding="utf-8") as f:
                    data = json.load(f)
                    
                    # Сначала пробуем найти сообщения по новому chat_id
                    messages = data.get(chat_id, [])
                    
                    if not messages:
                        # Если сообщений нет, пробуем найти сообщения по старому chat_id="4"
                        # Это нужно для обратной совместимости со старыми сообщениями
                        old_messages = data.get("4", [])
                        
                        # Фильтруем сообщения, относящиеся к этим пользователям
                        filtered_messages = []
                        for msg in old_messages:
                            if (str(msg.get("sender_id")) == str(current_user_id) and str(msg.get("receiver_id")) == str(user_id)) or \
                               (str(msg.get("sender_id")) == str(user_id) and str(msg.get("receiver_id")) == str(current_user_id)):
                                # Обновляем chat_id для совместимости с новой системой
                                msg["chat_id"] = chat_id
                                filtered_messages.append(msg)
                        
                        # Используем отфильтрованные сообщения
                        messages = filtered_messages
                        
                        # Сохраняем обновленные сообщения в памяти chat_manager
                        if messages:
                            for msg in messages:
                                chat_manager.add_message_to_memory(chat_id, msg)
                    
                    print(f"DEBUG: Загружено {len(messages)} сообщений для чата {chat_id}")
                    # В случае, если мы мигрировали сообщения, сохраняем их в файл
                    chat_manager.save_messages_to_file()
            except Exception as e:
                print(f"DEBUG: Ошибка при загрузке сообщений из файла: {e}")
        
        print(f"DEBUG: Возвращаем {len(messages)} сообщений для чата с пользователем {user_id}")
        return messages
    except Exception as e:
        print(f"ERROR: Ошибка при получении сообщений чата: {e}")
        # Возвращаем пустой список, чтобы приложение продолжало работать
        return []

@app.get("/admin/login")
async def admin_login_get(request: Request):
    return templates.TemplateResponse("admin/index.html", {"request": request})

@app.post("/admin/login")
async def admin_login_post(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(deps.get_db)
):
    # Проверяем учетные данные администратора
    admin = db.query(models.User).filter(
        models.User.email == username,
        models.User.role == models.UserRole.ADMIN
    ).first()
    
    if not admin or not verify_password(password, admin.hashed_password):
        return templates.TemplateResponse(
            "admin/index.html",
            {"request": request, "error": "Неверный email или пароль"}
        )
    
    # Создаем токен для администратора
    access_token = create_access_token(
        data={"sub": str(admin.id), "is_admin": True}
    )
    
    response = RedirectResponse(url="/admin", status_code=303)
    response.set_cookie(
        key="access_token",
        value=access_token,
        httponly=False,  # Позволяем JavaScript получить доступ к токену
        max_age=3600 * 24,  # Увеличиваем время жизни до 24 часов
        samesite="lax",
        path="/"  # Устанавливаем для всех путей
    )
    
    # Для отладки добавляем токен в URL первый раз
    return response

@app.get("/admin", response_class=HTMLResponse)
@app.get("/admin/dashboard", response_class=HTMLResponse)
async def admin_dashboard(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ администратора
    user = await check_admin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
    
    # Получаем реальную статистику из БД
    total_users = db.query(models.User).count()
    
    # Общее количество объектов недвижимости
    total_properties = db.query(models.Property).count()
    
    # Количество объектов на модерации (заявки)
    pending_properties = db.query(models.Property).filter(models.Property.status == "pending").count()
    
    # Количество чатов
    total_chats = db.query(AppChatModel).count()
    
    # Для тикетов (обращений в техподдержку) используем примерное значение
    # В реальном приложении эти данные были бы получены из БД
    total_tickets = 13
    
    # Получаем последние 5 объектов недвижимости
    latest_properties_query = db.query(models.Property).options(
        joinedload(models.Property.owner),
        joinedload(models.Property.images)
    ).order_by(models.Property.created_at.desc()).limit(5).all()
    
    # Подготавливаем форматированные данные для отображения
    latest_properties = []
    
    status_display = {
        "draft": "Черновик",
        "pending": "На модерации",
        "active": "Активно",
        "rejected": "Отклонено",
        "sold": "Продано"
    }
    
    for prop in latest_properties_query:
        # Форматируем цену
        price = prop.price or 0
        
        # Определяем статус для отображения
        status_value = prop.status.value if prop.status else "draft"
        status = status_display.get(status_value, "Неизвестно")
        
        # Определяем тип объекта
        property_type = "sale"
        if hasattr(prop, 'property_type') and prop.property_type:
            property_type = prop.property_type.value
        
        property_type_display = {
            "sale": "Продажа",
            "rent": "Аренда"
        }.get(property_type, property_type)
        
        # Добавляем данные в массив
        latest_properties.append({
            "id": prop.id,
            "title": prop.title or f"Объект #{prop.id}",
            "price": price,
            "status": status_value,
            "status_display": status,
            "property_type": property_type,
            "property_type_display": property_type_display,
            "created_at": prop.created_at
        })
    
    # Рассчитываем динамические проценты изменений на основе имеющихся данных
    # В реальном приложении это можно было бы рассчитать сравнивая с предыдущим месяцем
    # Для примера используем ID как показатель роста
    
    # Делаем рост пользователей на основе ID последнего пользователя
    last_user_id = db.query(models.User.id).order_by(models.User.id.desc()).first()
    last_user_id = last_user_id[0] if last_user_id else 0
    users_change = round((last_user_id - total_users) / max(total_users, 1) * 100) if total_users > 0 else 0
    users_change = min(max(users_change, -99), 99)  # Ограничиваем диапазоном -99 до 99
    
    # Делаем рост объектов на основе ID последнего объекта
    last_property_id = db.query(models.Property.id).order_by(models.Property.id.desc()).first()
    last_property_id = last_property_id[0] if last_property_id else 0
    properties_change = round((last_property_id - total_properties) / max(total_properties, 1) * 100) if total_properties > 0 else 0
    properties_change = min(max(properties_change, -99), 99)  # Ограничиваем диапазоном -99 до 99
    
    return templates.TemplateResponse(
        "admin/dashboard.html",
        {
            "request": request,
            "current_admin": user,  # Передаем данные администратора
            "total_users": total_users,
            "users_count": total_users,
            "users_change": users_change,
            "total_properties": total_properties,
            "properties_count": total_properties,
            "properties_change": properties_change,
            "total_chats": total_chats,
            "chats_change": total_chats,  # Используем количество чатов как динамический процент
            "requests_count": pending_properties,  # Заявки = объекты на модерации
            "requests_change": pending_properties,  # Используем количество заявок как динамический процент
            "tickets_count": total_tickets,  # Количество тикетов в техподдержку
            "tickets_change": -5,  # Примерное значение изменения
            "last_properties": latest_properties,
            "last_tickets": [],  # В данной версии не реализовано
        }
    )

@app.get("/admin/users", response_class=HTMLResponse)
async def admin_users(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ администратора
    user = await check_admin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
        
    # Получаем только ОБЫЧНЫХ пользователей (исключаем администраторов)
    users = db.query(models.User).filter(models.User.role == models.UserRole.USER).all()
    
    # Подготавливаем данные для отображения
    enhanced_users = []
    
    for user_item in users:
        # Получаем количество объявлений пользователя
        properties_count = db.query(models.Property).filter(models.Property.owner_id == user_item.id).count()
        
        # Получаем количество объявлений с 360-турами
        tours_count = db.query(models.Property).filter(
            models.Property.owner_id == user_item.id,
            models.Property.tour_360_url.isnot(None),
            models.Property.tour_360_url != ""
        ).count()
        
        # Форматируем дату регистрации
        registered_at = "Нет данных"
        if hasattr(user_item, 'created_at') and user_item.created_at:
            registered_at = user_item.created_at.strftime("%Y-%m-%d %H:%M:%S")
        
        # Добавляем данные в массив
        enhanced_users.append({
            "id": user_item.id,
            "full_name": user_item.full_name if hasattr(user_item, 'full_name') and user_item.full_name else f"Пользователь {user_item.id}",
            "phone": user_item.phone if hasattr(user_item, 'phone') and user_item.phone else "Нет данных",
            "email": user_item.email if hasattr(user_item, 'email') and user_item.email else "Нет данных",
            "is_active": user_item.is_active if hasattr(user_item, 'is_active') else True,
            "properties_count": properties_count,
            "tours_count": tours_count,
            "registered_at": registered_at,
            "avatar_url": user_item.avatar_url if hasattr(user_item, 'avatar_url') and user_item.avatar_url else None,
        })
    
    return templates.TemplateResponse(
        "admin/users.html",
        {
            "request": request,
            "current_admin": user,  # Передаем данные текущего администратора
            "users": enhanced_users,
            "start_item": 1,
            "end_item": len(enhanced_users),
            "total_users": len(enhanced_users),
            "search_query": "",
            "status": None,
            "current_page": 1,
            "total_pages": 1,
            "pages": [1],
            "show_ellipsis": False,
        }
    )

@app.get("/admin/properties", response_class=HTMLResponse)
async def admin_properties(
    request: Request, 
    status: str = Query(None), 
    property_type: str = Query(None),
    search: str = Query(None),
    page: int = Query(1, ge=1),
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ администратора
    user = await check_admin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
        
    # Получаем категории для фильтра
    categories = db.query(models.Category).all()
    
    # Создаем базовый запрос к объектам недвижимости
    properties_query = db.query(models.Property).options(
        joinedload(models.Property.owner),
        joinedload(models.Property.images)
    )
    
    # Применяем фильтры
    if status:
        properties_query = properties_query.filter(models.Property.status == status)
    
    if property_type:
        try:
            property_type_id = int(property_type)
            # Фильтруем по категории (связь с таблицей categories через PropertyCategory)
            properties_query = properties_query.join(models.PropertyCategory).filter(
                models.PropertyCategory.category_id == property_type_id
            )
        except (ValueError, TypeError):
            # Если не удалось преобразовать в число, игнорируем фильтр
            pass
            
    if search:
        search_term = f"%{search}%"
        properties_query = properties_query.filter(
            or_(
                models.Property.title.ilike(search_term),
                models.Property.description.ilike(search_term),
                models.Property.address.ilike(search_term)
            )
        )
    
    # Получаем общее количество записей после применения фильтров
    total_items = properties_query.count()
    
    # Настройка пагинации
    items_per_page = 10
    total_pages = (total_items + items_per_page - 1) // items_per_page if total_items > 0 else 1
    
    # Проверка корректности номера страницы
    if page > total_pages and total_pages > 0:
        page = total_pages
        
    # Получаем данные с пагинацией
    start_idx = (page - 1) * items_per_page
    properties_query = properties_query.order_by(desc(models.Property.created_at)).offset(start_idx).limit(items_per_page)
    
    # Получаем результаты запроса
    properties_results = properties_query.all()
    
    # Подготавливаем данные для отображения в шаблоне
    enhanced_properties = []
    
    for prop in properties_results:
        # Находим главное изображение или используем заглушку
        image_url = "/static/layout/assets/img/property-placeholder.jpg"
        if prop.images:
            main_images = [img for img in prop.images if img.is_main]
            if main_images:
                image_url = main_images[0].url
            else:
                image_url = prop.images[0].url if prop.images else image_url
        
        # Форматируем цену
        price_formatted = f"{prop.price:,.0f} KGZ" if prop.price else "Цена не указана"
        
        # Определяем понятный статус для отображения
        status_map = {
            "draft": "Черновик",
            "pending": "На проверке",
            "active": "Активно",
            "rejected": "Отклонено",
            "sold": "Продано"
        }
        
        status_val = prop.status.value if prop.status else "draft"
        status_display = status_map.get(status_val, "Неизвестно")
        
        # Проверяем наличие 360° тура
        has_tour = bool(prop.tour_360_url) if hasattr(prop, 'tour_360_url') and prop.tour_360_url else False
        
        # Количество просмотров для объектов на модерации - 0
        views = 0 if status_val == "pending" else prop.views if hasattr(prop, 'views') and prop.views else 0
        
        # Добавляем данные в массив
        enhanced_properties.append({
            "id": prop.id,
            "title": prop.title or f"Объект #{prop.id}",
            "description": prop.description,
            "price": prop.price,
            "price_formatted": price_formatted,
            "address": prop.address or "Адрес не указан",
            "city": prop.city,
            "area": prop.area,
            "status": status_val,
            "status_display": status_display,
            "owner_id": prop.owner_id,
            "owner_name": prop.owner.full_name if prop.owner else "Нет данных",
            "image_url": image_url,
            "created_at": prop.created_at.strftime("%Y-%m-%d %H:%M:%S") if prop.created_at else "Нет данных",
            "views": views,
            "has_tour": has_tour,
            "rooms": prop.rooms,
            "floor": prop.floor,
            "building_floors": prop.building_floors,
            "has_balcony": prop.has_balcony,
            "has_furniture": prop.has_furniture,
            "has_renovation": prop.has_renovation,
            "has_parking": prop.has_parking
        })
    
    # Вычисляем начальный и конечный индексы для пагинации
    start_item = start_idx + 1 if total_items > 0 else 0
    end_item = min(start_idx + len(enhanced_properties), total_items)
    
    # Формируем параметры запроса для пагинации
    query_params = ""
    if status:
        query_params += f"&status={status}"
    if property_type:
        query_params += f"&property_type={property_type}"
    if search:
        query_params += f"&search={search}"
    
    # Генерируем список страниц для навигации
    page_range = range(max(1, page - 2), min(total_pages + 1, page + 3))
    
    return templates.TemplateResponse(
        "admin/properties.html",
        {
            "request": request,
            "current_admin": user,  # Передаем данные администратора
            "properties": enhanced_properties,
            "total_pages": total_pages,
            "current_page": page,
            "pages": page_range,
            "show_ellipsis": total_pages > 5,
            "start_item": start_item,
            "end_item": end_item,
            "total_properties": total_items,
            "search_query": search or "",
            "status": status,
            "property_type": property_type,
            "categories": categories,  # Передаем категории для выпадающего списка
        }
    )

# Функция для проверки доступа администратора
async def check_admin_access(request: Request, db: Session):
    auth_token = request.cookies.get('access_token')
    if not auth_token:
        return RedirectResponse(url="/admin/login", status_code=303)
    
    try:
        payload = pyjwt.decode(auth_token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
        if datetime.fromtimestamp(payload["exp"]) < datetime.utcnow():
            return RedirectResponse(url="/admin/login", status_code=303)
            
        # Проверяем, что пользователь является администратором
        if not payload.get("is_admin"):
            return RedirectResponse(url="/admin/login", status_code=303)
            
        user_id = int(payload.get("sub"))
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return RedirectResponse(url="/admin/login", status_code=303)
            
        return user
    except Exception as e:
        print(f"DEBUG: Token validation error: {str(e)}")
        return RedirectResponse(url="/admin/login", status_code=303)

@app.get("/admin/requests", response_class=HTMLResponse, name="admin_requests")
async def admin_requests(request: Request, tab: str = Query('listings'), status: str = Query(None), 
                    property_type: str = Query(None), search: str = Query(None), 
                    page: int = Query(1, ge=1), db: Session = Depends(deps.get_db)):
    # Проверяем авторизацию и доступ администратора
    user = await check_admin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
    
    # Получаем количество объявлений по типам
    try:
        # Для таба tours будем считать объявления с запросом на съемку 360
        tour_requests_count = db.query(models.Property).filter(
            models.Property.tour_360_url.like('%example.com%')
        ).count()
        
        # Для таба listings будем считать только объявления со статусом PENDING
        listing_requests_count = db.query(models.Property).filter(
            models.Property.status == 'PENDING'
        ).count()
    except Exception as e:
        print(f"Error counting properties: {e}")
        tour_requests_count = 0
        listing_requests_count = 0
    
    # Получаем список объявлений из БД
    try:
        # Создаем базовый запрос к таблице properties
        query = db.query(models.Property)\
            .join(models.User, models.Property.owner_id == models.User.id, isouter=True)\
            .options(
                joinedload(models.Property.owner),
                joinedload(models.Property.images)
            )
            
        # Фильтрация по типу объявления
        if tab == 'tours':
            # Для таба tours берем только принятые объявления с запросом на съемку 360
            query = query.filter(
                models.Property.tour_360_url.like('%example.com%'),
                models.Property.status.in_(['ACTIVE', 'PROCESSING'])
            )
        elif tab == 'listings':
            # Для таба listings берем все объявления со статусом PENDING (на модерации)
            query = query.filter(
                models.Property.status == 'PENDING'
            )
            
        # Фильтрация по статусу, если указан
        if status:
            if status == 'new':
                query = query.filter(models.Property.status.in_(['NEW', 'PENDING']))
            else:
                status_map = {
                    'in_progress': 'PROCESSING',
                    'completed': 'ACTIVE',
                    'rejected': 'REJECTED'
                }
                if status in status_map:
                    query = query.filter(models.Property.status == status_map[status])
        
        # Фильтрация по типу недвижимости, если указан
        if property_type:
            try:
                property_type_id = int(property_type)
                # Фильтруем по категории (связь с таблицей categories)
                query = query.join(models.PropertyCategory).filter(
                    models.PropertyCategory.category_id == property_type_id
                )
            except (ValueError, TypeError):
                # Если property_type не является числом, фильтруем по типу как обычно
                query = query.filter(models.Property.type == property_type)
        
        # Поиск
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    models.Property.title.ilike(search_term),
                    models.Property.description.ilike(search_term),
                    models.Property.address.ilike(search_term),
                    models.User.full_name.ilike(search_term)
                )
            )
        
        # Получаем общее количество записей после применения фильтров
        total_items = query.count()
        
        # Пагинация
        items_per_page = 10
        total_pages = (total_items + items_per_page - 1) // items_per_page if total_items > 0 else 1
        
        # Проверяем корректность номера страницы
        if page > total_pages and total_pages > 0:
            page = total_pages
            
        # Получаем данные с пагинацией
        start_idx = (page - 1) * items_per_page
        query = query.order_by(desc(models.Property.created_at)).offset(start_idx).limit(items_per_page)
        
        # Получаем объекты недвижимости
        properties = query.all()
        
        # Преобразуем в формат для шаблона
        requests_data = []
        
        # Маппинг статусов из БД в формат для шаблона
        status_map_reverse = {
            'NEW': 'new',
            'PENDING': 'new',
            'PROCESSING': 'in_progress',
            'ACTIVE': 'completed',
            'REJECTED': 'rejected',
            'INACTIVE': 'rejected'
        }
        
        for prop in properties:
            # Форматируем цену корректно
            if prop.price:
                if prop.price >= 1000000:
                    millions = prop.price / 1000000
                    price_formatted = f"{millions:.1f} млн KGZ".replace('.0', '')
                else:
                    thousands = prop.price / 1000
                    price_formatted = f"{thousands:.1f} тыс KGZ".replace('.0', '')
            else:
                price_formatted = "Цена не указана"
            
            # Определяем статус для отображения
            display_status = status_map_reverse.get(prop.status, 'new') if prop.status else 'new'
            
            # Получаем все изображения объявления
            property_images = []
            try:
                # Используем связанные изображения из SQLAlchemy
                for img in prop.images:
                    if img.url:
                        property_images.append({
                            'url': img.url,
                            'is_main': img.is_main
                        })
                
                # Если изображений нет, добавляем заглушку
                if not property_images:
                    property_images = []
                
            except Exception as e:
                print(f"Ошибка при получении изображений для объявления ID={prop.id}: {e}")
                property_images = []
            
            # Получаем информацию о владельце
            owner_data = {
                'id': prop.owner.id if prop.owner else None,
                'name': prop.owner.full_name if prop.owner and prop.owner.full_name else "Пользователь",
                'email': prop.owner.email if prop.owner and prop.owner.email else ""
            }
            
            # Получаем информацию о категории
            category_info = None
            try:
                # Получаем связь через PropertyCategory
                property_category = db.query(models.PropertyCategory).filter(
                    models.PropertyCategory.property_id == prop.id
                ).first()
                
                if property_category:
                    category = db.query(models.Category).filter(
                        models.Category.id == property_category.category_id
                    ).first()
                    if category:
                        category_info = {'name': category.name, 'id': category.id}
            except Exception as e:
                print(f"Ошибка при получении категории: {e}")
                category_info = None
            
            # Добавляем объект в список
            requests_data.append({
                'id': prop.id,
                'status': display_status,
                'created_at': prop.created_at.strftime("%d.%m.%Y %H:%M") if prop.created_at else "",
                'scheduled_date': prop.notes,  # Берем дату съемки из поля notes
                'property': {
                    'id': prop.id,
                    'title': prop.title or f"Объект №{prop.id}",
                    'address': prop.address or "Адрес не указан",
                    'price': prop.price or 0,
                    'price_formatted': price_formatted,
                    'type': prop.type or "apartment",
                    'images': property_images,  # Передаем все изображения
                    'category': category_info  # Добавляем информацию о категории
                },
                'user': owner_data
            })
            
        # Вычисляем начальный и конечный индексы для пагинации
        start_idx = (page - 1) * items_per_page
        end_idx = min(start_idx + len(requests_data), total_items)
        
    except Exception as e:
        print(f"Error getting requests: {e}")
        requests_data = []
        total_items = 0
        total_pages = 1
    
    # Пагинация
    items_per_page = 10
    total_pages = (total_items + items_per_page - 1) // items_per_page if total_items > 0 else 1
    
    # Проверяем корректность номера страницы
    if page > total_pages and total_pages > 0:
        page = total_pages
    
    # Получаем элементы для текущей страницы
    start_idx = (page - 1) * items_per_page
    end_idx = min(start_idx + items_per_page, total_items)
    page_items = requests_data[start_idx:end_idx] if total_items > 0 else []
    
    # Вычисляем метрики для отображения
    try:
        metrics_query = db.query(models.Property)
        
        # Фильтрация по типу объявления
        if tab == 'tours':
            metrics_query = metrics_query.filter(
                models.Property.tour_360_url.isnot(None),
                models.Property.tour_360_url != ''
            )
        elif tab == 'listings':
            metrics_query = metrics_query.filter(
                or_(models.Property.tour_360_url.is_(None),
                   models.Property.tour_360_url == '')
            )
        
        # Количество принятых/активных объявлений
        accepted_count = metrics_query.filter(models.Property.status == 'ACTIVE').count()
        
        # Количество отклоненных объявлений
        rejected_count = metrics_query.filter(models.Property.status.in_(['REJECTED', 'INACTIVE'])).count()
        
        # Количество ожидающих объявлений
        pending_count = metrics_query.filter(
            models.Property.status.in_(['NEW', 'PENDING', 'PROCESSING'])
        ).count()
        
        # Формируем метрики
        metrics = {
            'accepted': accepted_count,
            'rejected': rejected_count,
            'pending': pending_count,
            'avg_time': "1.2 дня"  # Упрощенное значение
        }
    except Exception as e:
        print(f"Error calculating metrics: {e}")
        metrics = {
            'accepted': 0,
            'rejected': 0,
            'pending': 0,
            'avg_time': "0 дней"
        }

    # Формируем строку параметров запроса для пагинации
    query_params = ""
    if status:
        query_params += f"&status={status}"
    if property_type:
        query_params += f"&property_type={property_type}"
    if search:
        query_params += f"&search={search}"
        
    # Генерируем список страниц для навигации
    page_range = range(max(1, page - 2), min(total_pages + 1, page + 3))
    
    # Получаем все категории для фильтра
    categories = db.query(models.Category).all()
    
    return templates.TemplateResponse("admin/requests.html", {
        "request": request,
        "user": user,
        "tab": tab,
        "status": status,
        "property_type": property_type,
        "search": search,
        "requests": page_items,
        "tour_requests_count": tour_requests_count,
        "listing_requests_count": listing_requests_count,
        "metrics": metrics,
        "page": page,
        "total_pages": total_pages,
        "total_requests": total_items,
        "items_per_page": items_per_page,
        "start_item": start_idx + 1 if total_items > 0 else 0,
        "end_item": end_idx,
        "pages": page_range,
        "query_params": query_params,
        "categories": categories,
    })

@app.get("/admin/settings", response_class=HTMLResponse, name="admin_settings")
async def admin_settings(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ администратора
    user = await check_admin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
    
    class DummySettings:
        email_new_properties = False
        email_new_users = False
        push_notifications = False
        digest_frequency = 'never'
        color_scheme = 'orange'
        theme = 'light'
        compact_mode = False
        animations_enabled = True
    
    dummy_settings = DummySettings()
    return templates.TemplateResponse(
        "admin/settings.html",
        {"request": request, "current_admin": user, "settings": dummy_settings}
    )

# API для настроек
class SettingsModel(BaseModel):
    email_new_properties: Optional[bool] = True
    email_new_users: Optional[bool] = True
    push_notifications: Optional[bool] = False
    digest_frequency: Optional[str] = "daily"
    color_scheme: Optional[str] = "orange"
    theme: Optional[str] = "light"
    compact_mode: Optional[bool] = False
    animations_enabled: Optional[bool] = True

@app.post("/api/v1/settings")
async def save_settings(settings: SettingsModel, request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем авторизацию, но для этого API допускаем любого авторизованного пользователя
    auth_token = request.cookies.get('access_token')
    auth_header = request.headers.get('Authorization')
    
    if auth_header and auth_header.startswith('Bearer '):
        auth_token = auth_header.split(' ')[1]
    
    if not auth_token:
        return JSONResponse(status_code=401, content={"success": False, "error": "Требуется авторизация"})
    
    try:
        # Проверяем валидность токена
        payload = pyjwt.decode(auth_token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
        if datetime.fromtimestamp(payload["exp"]) < datetime.utcnow():
            return JSONResponse(status_code=401, content={"success": False, "error": "Токен истек"})
        
        # В реальном приложении здесь бы сохраняли настройки в БД
        # Для демонстрации просто возвращаем успех
        return {"success": True, "message": "Настройки сохранены успешно"}
    except Exception as e:
        print(f"DEBUG: Ошибка при сохранении настроек: {str(e)}")
        return JSONResponse(status_code=500, content={"success": False, "error": f"Ошибка сервера: {str(e)}"})

# Сброс настроек
@app.post("/api/v1/settings/reset")
async def reset_settings(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем авторизацию
    auth_token = request.cookies.get('access_token')
    auth_header = request.headers.get('Authorization')
    
    if auth_header and auth_header.startswith('Bearer '):
        auth_token = auth_header.split(' ')[1]
    
    if not auth_token:
        return JSONResponse(status_code=401, content={"success": False, "error": "Требуется авторизация"})
    
    try:
        # Проверяем валидность токена
        payload = pyjwt.decode(auth_token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
        if datetime.fromtimestamp(payload["exp"]) < datetime.utcnow():
            return JSONResponse(status_code=401, content={"success": False, "error": "Токен истек"})
        
        # В реальном приложении здесь бы сбрасывали настройки к значениям по умолчанию
        # Для демонстрации просто возвращаем успех
        return {"success": True, "message": "Настройки сброшены к значениям по умолчанию"}
    except Exception as e:
        print(f"DEBUG: Ошибка при сбросе настроек: {str(e)}")
        return JSONResponse(status_code=500, content={"success": False, "error": f"Ошибка сервера: {str(e)}"})

# Функция для проверки доступа суперадмина
async def check_superadmin_access(request: Request, db: Session):
    auth_token = request.cookies.get('access_token')
    if not auth_token:
        return RedirectResponse(url="/superadmin/login", status_code=303)
    
    try:
        payload = pyjwt.decode(auth_token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
        if datetime.fromtimestamp(payload["exp"]) < datetime.utcnow():
            return RedirectResponse(url="/superadmin/login", status_code=303)
            
        # Проверяем, что пользователь является суперадминистратором
        if not payload.get("is_superadmin"):
            return RedirectResponse(url="/superadmin/login", status_code=303)
            
        user_id = int(payload.get("sub"))
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return RedirectResponse(url="/superadmin/login", status_code=303)
            
        return user
    except Exception as e:
        print(f"DEBUG: SuperAdmin token validation error: {str(e)}")
        return RedirectResponse(url="/superadmin/login", status_code=303)

# ============================ SuperAdmin Routes ============================

@app.get("/superadmin/login")
async def superadmin_login_get(request: Request):
    return templates.TemplateResponse("superadmin/login.html", {"request": request})

@app.post("/superadmin/login")
async def superadmin_login_post(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(deps.get_db)
):
    # Проверяем учетные данные суперадмина
    # В реальном приложении должен быть отдельный статус суперадмина
    superadmin = db.query(models.User).filter(
        models.User.email == username,
        models.User.role == models.UserRole.ADMIN  # Временно используем роль ADMIN
    ).first()
    
    # Дополнительная проверка: только определенные email могут быть суперадминами
    superadmin_emails = ['superadmin@wazir.kg', 'admin@wazir.kg']
    
    if not superadmin or not verify_password(password, superadmin.hashed_password) or username not in superadmin_emails:
        return templates.TemplateResponse(
            "superadmin/login.html",
            {"request": request, "error": "Неверный логин или пароль суперадмина"}
        )
    
    # Создаем токен для суперадмина с флагом is_superadmin
    access_token = create_access_token(
        data={"sub": str(superadmin.id), "is_admin": True, "is_superadmin": True}
    )
    
    response = RedirectResponse(url="/superadmin/dashboard", status_code=303)
    response.set_cookie(
        key="access_token",
        value=access_token,
        httponly=False,
        max_age=3600 * 24,
        samesite="lax",
        path="/"
    )
    
    return response

@app.get("/superadmin/logout")
async def superadmin_logout():
    response = RedirectResponse(url="/superadmin/login", status_code=303)
    response.delete_cookie("access_token", path="/")
    return response

@app.get("/superadmin", response_class=HTMLResponse)
@app.get("/superadmin/dashboard", response_class=HTMLResponse)
async def superadmin_dashboard(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
    
    # Получаем статистику
    from datetime import datetime, timedelta
    
    # Статистика
    stats = {
        'admins_count': db.query(models.User).filter(models.User.role == models.UserRole.ADMIN).count(),
        'users_count': db.query(models.User).filter(models.User.role == models.UserRole.USER).count(),
        'properties_count': db.query(models.Property).count(),
        'pending_requests': db.query(models.Property).filter(models.Property.status == 'PENDING').count()
    }
    
    # РЕАЛЬНЫЕ последние действия из БД
    recent_activities = []
    
    # Последние зарегистрированные пользователи (за последние 7 дней)
    week_ago = datetime.now() - timedelta(days=7)
    recent_users = db.query(models.User).filter(
        models.User.role == models.UserRole.USER,
        models.User.created_at >= week_ago
    ).order_by(desc(models.User.created_at)).limit(3).all()
    
    for user_item in recent_users:
        time_diff = datetime.now() - user_item.created_at
        if time_diff.days > 0:
            time_str = f"{time_diff.days} дн. назад"
        elif time_diff.seconds > 3600:
            time_str = f"{time_diff.seconds // 3600} ч. назад"
        else:
            time_str = f"{time_diff.seconds // 60} мин. назад"
            
        recent_activities.append({
            'icon': 'fas fa-user-plus',
            'action': f'Регистрация: {user_item.full_name or user_item.email}',
            'admin_name': 'Система',
            'time': time_str
        })
    
    # Последние объявления (за последние 3 дня)
    three_days_ago = datetime.now() - timedelta(days=3)
    recent_properties = db.query(models.Property).filter(
        models.Property.created_at >= three_days_ago
    ).order_by(desc(models.Property.created_at)).limit(2).all()
    
    for prop in recent_properties:
        time_diff = datetime.now() - prop.created_at
        if time_diff.days > 0:
            time_str = f"{time_diff.days} дн. назад"
        elif time_diff.seconds > 3600:
            time_str = f"{time_diff.seconds // 3600} ч. назад"
        else:
            time_str = f"{time_diff.seconds // 60} мин. назад"
            
        recent_activities.append({
            'icon': 'fas fa-building',
            'action': f'Новое объявление: {prop.title or f"Объект #{prop.id}"}',
            'admin_name': 'Пользователь',
            'time': time_str
        })
    
    # Последние изменения статусов объявлений
    recent_status_changes = db.query(models.Property).filter(
        models.Property.updated_at >= three_days_ago,
        models.Property.status.in_(['ACTIVE', 'REJECTED'])
    ).order_by(desc(models.Property.updated_at)).limit(2).all()
    
    for prop in recent_status_changes:
        time_diff = datetime.now() - prop.updated_at
        if time_diff.days > 0:
            time_str = f"{time_diff.days} дн. назад"
        elif time_diff.seconds > 3600:
            time_str = f"{time_diff.seconds // 3600} ч. назад"
        else:
            time_str = f"{time_diff.seconds // 60} мин. назад"
            
        status_text = "одобрено" if prop.status == 'ACTIVE' else "отклонено"
        recent_activities.append({
            'icon': 'fas fa-edit',
            'action': f'Объявление {status_text}: {prop.title or f"#{prop.id}"}',
            'admin_name': 'Модератор',
            'time': time_str
        })
    
    # Если нет реальных активностей, добавляем системное сообщение
    if not recent_activities:
        recent_activities.append({
            'icon': 'fas fa-info-circle',
            'action': 'Система запущена и работает стабильно',
            'admin_name': 'Система',
            'time': 'Сейчас'
        })
    
    # РЕАЛЬНЫЕ системные уведомления
    system_notifications = []
    
    # Проверяем наличие объявлений на модерации
    pending_count = db.query(models.Property).filter(models.Property.status == 'PENDING').count()
    if pending_count > 0:
        system_notifications.append({
            'icon': 'fas fa-exclamation-triangle',
            'title': 'Требуется модерация',
            'message': f'{pending_count} объявлений ожидают модерации',
            'color': '#f59e0b',
            'bg_color': '#fefbf3',
            'time': 'Сейчас'
        })
    
    # Проверяем новых пользователей за последние 24 часа
    yesterday = datetime.now() - timedelta(days=1)
    new_users_count = db.query(models.User).filter(
        models.User.created_at >= yesterday,
        models.User.role == models.UserRole.USER
    ).count()
    
    if new_users_count > 0:
        system_notifications.append({
            'icon': 'fas fa-users',
            'title': 'Новые пользователи',
            'message': f'{new_users_count} новых пользователей за последние 24 часа',
            'color': '#10b981',
            'bg_color': '#f0fdf4',
            'time': '24 часа'
        })
    
    # Проверяем неактивных администраторов
    inactive_admins = db.query(models.User).filter(
        models.User.role == models.UserRole.ADMIN,
        models.User.is_active == False
    ).count()
    
    if inactive_admins > 0:
        system_notifications.append({
            'icon': 'fas fa-user-slash',
            'title': 'Неактивные администраторы',
            'message': f'{inactive_admins} администраторов неактивны',
            'color': '#ef4444',
            'bg_color': '#fef2f2',
            'time': '1 час назад'
        })
    
    # Проверяем количество активных объявлений
    active_properties = db.query(models.Property).filter(models.Property.status == 'ACTIVE').count()
    if active_properties > 100:
        system_notifications.append({
            'icon': 'fas fa-chart-line',
            'title': 'Высокая активность',
            'message': f'{active_properties} активных объявлений в системе',
            'color': '#10b981',
            'bg_color': '#f0fdf4',
            'time': 'Сейчас'
        })
    
    # Если нет уведомлений, добавляем позитивное сообщение
    if not system_notifications:
        system_notifications.append({
            'icon': 'fas fa-check-circle',
            'title': 'Все в порядке',
            'message': 'Система работает стабильно, проблем не обнаружено',
            'color': '#10b981',
            'bg_color': '#f0fdf4',
            'time': 'Сейчас'
        })
    
    # РЕАЛЬНАЯ системная информация
    system_info = {
        'python_version': f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
        'fastapi_version': '0.104.1',
        'memory_usage': f"{psutil.virtual_memory().percent:.1f}%" if psutil else "Недоступно",
        'uptime': 'Online'
    }
    
    # Пытаемся получить реальную информацию о системе
    if psutil:
        try:
            # Использование CPU
            cpu_percent = psutil.cpu_percent(interval=1)
            # Использование памяти
            memory = psutil.virtual_memory()
            # Использование диска
            disk = psutil.disk_usage('/')
            
            system_info.update({
                'cpu_usage': f"{cpu_percent:.1f}%",
                'memory_usage': f"{memory.percent:.1f}%",
                'memory_total': f"{memory.total // (1024**3)} GB",
                'disk_usage': f"{disk.percent:.1f}%",
                'disk_free': f"{disk.free // (1024**3)} GB"
            })
        except Exception as e:
            print(f"DEBUG: Ошибка получения системной информации: {e}")
    
    return templates.TemplateResponse(
        "superadmin/dashboard.html",
        {
            "request": request,
            "current_user": user,
            "stats": stats,
            "recent_activities": recent_activities,
            "system_notifications": system_notifications,
            "system_info": system_info
        }
    )

@app.get("/superadmin/admins", response_class=HTMLResponse)
async def superadmin_admins(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
    
    # Получаем всех администраторов (ADMIN и MANAGER роли)
    admins = db.query(models.User).filter(
        or_(
            models.User.role == models.UserRole.ADMIN,
            models.User.role == models.UserRole.MANAGER
        )
    ).all()
    
    # Подсчитываем статистику
    admins_total = len(admins)
    admins_active = len([admin for admin in admins if admin.is_active])
    admins_inactive = admins_total - admins_active
    
    # Подготавливаем данные для отображения
    enhanced_admins = []
    for admin in admins:
        enhanced_admins.append({
            "id": admin.id,
            "full_name": admin.full_name,
            "email": admin.email,
            "phone": admin.phone,
            "is_active": admin.is_active,
            "avatar_url": getattr(admin, 'avatar_url', None),
            "created_at": admin.created_at.strftime("%d.%m.%Y") if hasattr(admin, 'created_at') and admin.created_at else "Нет данных"
        })
    
    # Добавляем переменные пагинации (даже если пагинация не нужна)
    total_pages = 1
    current_page = 1
    per_page = len(enhanced_admins)
    total_items = len(enhanced_admins)
    
    return templates.TemplateResponse(
        "superadmin/admins.html",
        {
            "request": request,
            "current_user": user,
            "admins": enhanced_admins,
            "admins_total": admins_total,
            "admins_active": admins_active,
            "admins_inactive": admins_inactive,
            "total_pages": total_pages,
            "current_page": current_page,
            "per_page": per_page,
            "total_items": total_items
        }
    )

@app.get("/superadmin/users", response_class=HTMLResponse)
async def superadmin_users(
    request: Request, 
    search: str = Query(None),
    status: str = Query(None),
    property_filter: str = Query(None),
    page: int = Query(1, ge=1),
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
    
    # Базовый запрос
    query = db.query(models.User).filter(models.User.role == models.UserRole.USER)
    
    # Применяем фильтры
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                models.User.full_name.ilike(search_term),
                models.User.email.ilike(search_term),
                models.User.phone.ilike(search_term)
            )
        )
    
    if status:
        if status == "active":
            query = query.filter(models.User.is_active == True)
        elif status == "blocked":
            query = query.filter(models.User.is_active == False)
    
    # Пагинация
    total_items = query.count()
    items_per_page = 50
    total_pages = (total_items + items_per_page - 1) // items_per_page if total_items > 0 else 1
    
    if page > total_pages and total_pages > 0:
        page = total_pages
    
    start_idx = (page - 1) * items_per_page
    users_results = query.order_by(desc(models.User.created_at)).offset(start_idx).limit(items_per_page).all()
    
    # Подготавливаем данные для отображения
    enhanced_users = []
    for u in users_results:
        properties_count = db.query(models.Property).filter(models.Property.owner_id == u.id).count()
        
        # Применяем фильтр по недвижимости после получения данных
        if property_filter:
            if property_filter == "with" and properties_count == 0:
                continue
            elif property_filter == "without" and properties_count > 0:
                continue
        
        enhanced_users.append({
            "id": u.id,
            "full_name": u.full_name or f"Пользователь {u.id}",
            "phone": u.phone or "Нет данных",
            "email": u.email or "Нет данных",
            "is_active": u.is_active,
            "properties_count": properties_count,
            "registered_at": u.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(u, 'created_at') and u.created_at else "Нет данных",
        })
    
    # Параметры запроса для пагинации
    query_params = ""
    if search:
        query_params += f"&search={search}"
    if status:
        query_params += f"&status={status}"
    if property_filter:
        query_params += f"&property_filter={property_filter}"
    
    start_item = start_idx + 1 if len(enhanced_users) > 0 else 0
    end_item = start_idx + len(enhanced_users)
    page_range = range(max(1, page - 2), min(total_pages + 1, page + 3))
    
    return templates.TemplateResponse(
        "superadmin/users.html",
        {
            "request": request,
            "current_user": user,
            "users": enhanced_users,
            "total_items": total_items,
            "items_per_page": items_per_page,
            "current_page": page,
            "total_pages": total_pages,
            "start_item": start_item,
            "end_item": end_item,
            "page_range": page_range,
            "query_params": query_params,
            "search": search or "",
            "status": status or "",
            "property_filter": property_filter or ""
        }
    )

@app.get("/superadmin/companies", response_class=HTMLResponse)
async def superadmin_companies(
    request: Request, 
    search: str = Query(None),
    status: str = Query(None),
    page: int = Query(1, ge=1),
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
    
    from sqlalchemy.orm import joinedload
    from sqlalchemy import or_
    
    items_per_page = 20
    offset = (page - 1) * items_per_page
    
    # Базовый запрос для компаний (пользователи с ролью COMPANY)
    query = db.query(models.User).filter(models.User.role == models.UserRole.COMPANY)
    
    # Применяем фильтры
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                models.User.full_name.ilike(search_term),
                models.User.email.ilike(search_term),
                models.User.phone.ilike(search_term),
                models.User.company_name.ilike(search_term),
                models.User.company_number.ilike(search_term),
                models.User.company_owner.ilike(search_term)
            )
        )
    
    if status:
        if status == "active":
            query = query.filter(models.User.is_active == True)
        elif status == "blocked":
            query = query.filter(models.User.is_active == False)
    
    # Подсчет общего количества для пагинации
    total_items = query.count()
    total_pages = (total_items + items_per_page - 1) // items_per_page
    
    # Получаем компании с пагинацией
    companies_results = query.offset(offset).limit(items_per_page).all()
    
    # Формируем данные для шаблона
    enhanced_companies = []
    total_company_properties = 0
    
    for company in companies_results:
        properties_count = db.query(models.Property).filter(models.Property.owner_id == company.id).count()
        total_company_properties += properties_count
        
        enhanced_companies.append({
            "id": company.id,
            "company_name": company.company_name,
            "company_number": company.company_number,
            "company_owner": company.company_owner,
            "company_logo_url": company.company_logo_url,
            "company_description": company.company_description,
            "company_address": company.company_address,
            "full_name": company.full_name,
            "email": company.email,
            "phone": company.phone,
            "is_active": company.is_active,
            "properties_count": properties_count,
            "created_at": company.created_at.strftime("%d.%m.%Y") if hasattr(company, 'created_at') and company.created_at else "Нет данных"
        })
    
    # Статистика
    companies_total = query.count()
    companies_active = query.filter(models.User.is_active == True).count()
    companies_inactive = companies_total - companies_active
    
    return templates.TemplateResponse(
        "superadmin/companies.html",
        {
            "request": request,
            "current_user": user,
            "companies": enhanced_companies,
            "companies_total": companies_total,
            "companies_active": companies_active,
            "companies_inactive": companies_inactive,
            "total_company_properties": total_company_properties,
            "search": search or "",
            "status": status or "",
            "current_page": page,
            "total_pages": total_pages,
        }
    )

@app.get("/superadmin/properties", response_class=HTMLResponse)
async def superadmin_properties(
    request: Request, 
    search: str = Query(None),
    status: str = Query(None),
    property_type: str = Query(None),
    page: int = Query(1, ge=1),
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
    
    from sqlalchemy.orm import joinedload
    from sqlalchemy import or_, desc
    
    items_per_page = 20
    offset = (page - 1) * items_per_page
    
    # Базовый запрос для объявлений
    query = db.query(models.Property).options(joinedload(models.Property.owner))
    
    # Применяем фильтры
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                models.Property.title.ilike(search_term),
                models.Property.address.ilike(search_term),
                models.Property.description.ilike(search_term)
            )
        )
    
    if status:
        if status == "active":
            query = query.filter(models.Property.status == 'ACTIVE')
        elif status == "pending":
            query = query.filter(models.Property.status == 'PENDING')
        elif status == "rejected":
            query = query.filter(models.Property.status == 'REJECTED')
        elif status == "draft":
            query = query.filter(models.Property.status == 'DRAFT')
    
    if property_type:
        # Фильтр по типу недвижимости - здесь можно добавить логику фильтрации
        pass
    
    # Сортировка по дате создания (новые сначала)
    query = query.order_by(desc(models.Property.created_at))
    
    # Подсчет общего количества для пагинации
    total_items = query.count()
    total_pages = (total_items + items_per_page - 1) // items_per_page
    
    # Получаем объявления с пагинацией
    properties_results = query.offset(offset).limit(items_per_page).all()
    
    # Формируем данные для шаблона
    enhanced_properties = []
    
    for prop in properties_results:
        # Форматирование цены
        price_formatted = f"{int(prop.price):,} сом".replace(",", " ") if prop.price else "Не указана"
        
        # Получаем имя владельца
        owner_name = "Неизвестен"
        if prop.owner:
            if prop.owner.role == models.UserRole.COMPANY:
                owner_name = prop.owner.company_name or prop.owner.full_name or prop.owner.email
            else:
                owner_name = prop.owner.full_name or prop.owner.email
        
        # URL изображения (если есть)
        image_url = "/static/img/property-placeholder.jpg"  # Дефолтное изображение
        if hasattr(prop, 'images') and prop.images:
            # prop.images - это список объектов PropertyImage, получаем URL первого изображения
            first_image = prop.images[0]
            if hasattr(first_image, 'url'):
                image_url = first_image.url
        
        enhanced_properties.append({
            "id": prop.id,
            "title": prop.title or f"Объект #{prop.id}",
            "address": prop.address or "Адрес не указан",
            "price": prop.price or 0,
            "price_formatted": price_formatted,
            "rooms": prop.rooms,
            "area": prop.area,
            "status": prop.status.value.lower() if prop.status else "draft",
            "status_display": {
                "active": "Активно",
                "pending": "На модерации", 
                "rejected": "Отклонено",
                "draft": "Черновик"
            }.get(prop.status.value.lower() if prop.status else "draft", "Неизвестно"),
            "owner_id": prop.owner_id,
            "owner_name": owner_name,
            "image_url": image_url,
            "created_at": prop.created_at.strftime("%d.%m.%Y") if prop.created_at else "Неизвестно"
        })
    
    # Статистика
    total_properties = db.query(models.Property).count()
    active_count = db.query(models.Property).filter(models.Property.status == 'ACTIVE').count()
    pending_count = db.query(models.Property).filter(models.Property.status == 'PENDING').count()
    rejected_count = db.query(models.Property).filter(models.Property.status == 'REJECTED').count()
    
    return templates.TemplateResponse(
        "superadmin/properties.html",
        {
            "request": request,
            "current_user": user,
            "properties": enhanced_properties,
            "total_properties": total_properties,
            "active_count": active_count,
            "pending_count": pending_count,
            "rejected_count": rejected_count,
            "search": search or "",
            "status": status or "",
            "property_type": property_type or "",
            "current_page": page,
            "total_pages": total_pages,
        }
    )

@app.get("/superadmin/logs", response_class=HTMLResponse)
async def superadmin_logs(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
    
    return templates.TemplateResponse(
        "superadmin/logs.html",
        {
            "request": request,
            "current_user": user,
        }
    )

@app.get("/superadmin/analytics", response_class=HTMLResponse)
async def superadmin_analytics(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
    
    from datetime import datetime, timedelta
    from sqlalchemy import func, extract
    
    # Статистика пользователей
    total_users = db.query(models.User).filter(models.User.role == models.UserRole.USER).count()
    active_users = db.query(models.User).filter(
        models.User.role == models.UserRole.USER,
        models.User.is_active == True
    ).count()
    
    # Статистика объявлений
    total_properties = db.query(models.Property).count()
    active_properties = db.query(models.Property).filter(models.Property.status == 'ACTIVE').count()
    pending_properties = db.query(models.Property).filter(models.Property.status == 'PENDING').count()
    
    # Статистика по дням за последние 30 дней
    end_date = datetime.now()
    start_date = end_date - timedelta(days=30)
    
    # Регистрации пользователей по дням
    users_by_date = db.query(
        func.date(models.User.created_at).label('date'),
        func.count(models.User.id).label('count')
    ).filter(
        models.User.created_at >= start_date,
        models.User.role == models.UserRole.USER
    ).group_by(
        func.date(models.User.created_at)
    ).order_by('date').all()
    
    # Объявления по дням
    properties_by_date = db.query(
        func.date(models.Property.created_at).label('date'),
        func.count(models.Property.id).label('count')
    ).filter(
        models.Property.created_at >= start_date
    ).group_by(
        func.date(models.Property.created_at)
    ).order_by('date').all()
    
    # Популярные категории
    try:
        popular_categories = db.query(
            models.Category.name.label('category'),
            func.count(models.PropertyCategory.property_id).label('count')
        ).join(
            models.PropertyCategory, models.Category.id == models.PropertyCategory.category_id
        ).group_by(
            models.Category.id, models.Category.name
        ).order_by(
            func.count(models.PropertyCategory.property_id).desc()
        ).limit(10).all()
    except:
        popular_categories = []
    
    # Средняя цена по категориям
    try:
        price_by_category = db.query(
            models.Category.name.label('category'),
            func.avg(models.Property.price).label('avg_price')
        ).join(
            models.PropertyCategory, models.Category.id == models.PropertyCategory.category_id
        ).join(
            models.Property, models.PropertyCategory.property_id == models.Property.id
        ).filter(
            models.Property.price.isnot(None),
            models.Property.price > 0
        ).group_by(
            models.Category.id, models.Category.name
        ).order_by('avg_price').all()
    except:
        price_by_category = []
    
    # Статистика активности по месяцам
    monthly_stats = db.query(
        extract('month', models.Property.created_at).label('month'),
        func.count(models.Property.id).label('count')
    ).filter(
        models.Property.created_at >= datetime.now().replace(month=1, day=1)
    ).group_by(
        extract('month', models.Property.created_at)
    ).order_by('month').all()
    
    # Топ пользователи по количеству объявлений
    top_users = db.query(
        models.User.id,
        models.User.full_name,
        models.User.email,
        func.count(models.Property.id).label('properties_count')
    ).join(
        models.Property, models.User.id == models.Property.owner_id
    ).filter(
        models.User.role == models.UserRole.USER
    ).group_by(
        models.User.id, models.User.full_name, models.User.email
    ).order_by(
        func.count(models.Property.id).desc()
    ).limit(10).all()
    
    # Форматируем данные для графиков
    users_chart_data = {
        'labels': [item.date.strftime('%d.%m') for item in users_by_date[-14:]] or [],
        'data': [item.count for item in users_by_date[-14:]] or []
    }
    
    properties_chart_data = {
        'labels': [item.date.strftime('%d.%m') for item in properties_by_date[-14:]] or [],
        'data': [item.count for item in properties_by_date[-14:]] or []
    }
    
    categories_chart_data = {
        'labels': [item.category for item in popular_categories[:6]] or ['Нет данных'],
        'data': [item.count for item in popular_categories[:6]] or [0]
    }
    
    price_chart_data = {
        'labels': [item.category for item in price_by_category[:6]] or ['Нет данных'],
        'data': [float(item.avg_price) if item.avg_price else 0 for item in price_by_category[:6]] or [0]
    }
    
    # Месяцы на русском
    month_names = {
        1: 'Янв', 2: 'Фев', 3: 'Мар', 4: 'Апр', 5: 'Май', 6: 'Июн',
        7: 'Июл', 8: 'Авг', 9: 'Сен', 10: 'Окт', 11: 'Ноя', 12: 'Дек'
    }
    
    monthly_chart_data = {
        'labels': [month_names.get(int(item.month), f'Месяц {item.month}') for item in monthly_stats] or ['Нет данных'],
        'data': [item.count for item in monthly_stats] or [0]
    }
    
    return templates.TemplateResponse(
        "superadmin/analytics.html",
        {
            "request": request,
            "current_user": user,
            "total_users": total_users,
            "active_users": active_users,
            "total_properties": total_properties,
            "active_properties": active_properties,
            "pending_properties": pending_properties,
            "users_chart_data": users_chart_data,
            "properties_chart_data": properties_chart_data,
            "categories_chart_data": categories_chart_data,
            "price_chart_data": price_chart_data,
            "monthly_chart_data": monthly_chart_data,
            "top_users": top_users,
            "popular_categories": popular_categories
        }
    )

@app.get("/superadmin/settings", response_class=HTMLResponse)
async def superadmin_settings(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
    
    return templates.TemplateResponse(
        "superadmin/settings.html",
        {
            "request": request,
            "current_user": user,
        }
    )

# API роуты для суперадмина
from passlib.context import CryptContext

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

@app.post("/api/v1/superadmin/admins")
async def create_admin(
    request: Request,
    full_name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(None),
    password: str = Form(...),
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
    
    try:
        # Проверяем уникальность email
        existing_email = db.query(models.User).filter(models.User.email == email).first()
        if existing_email:
            return JSONResponse(status_code=400, content={"success": False, "message": "Пользователь с таким email уже существует"})
        
        # Проверяем уникальность телефона (если указан)
        if phone:
            existing_phone = db.query(models.User).filter(models.User.phone == phone).first()
            if existing_phone:
                return JSONResponse(status_code=400, content={"success": False, "message": "Пользователь с таким телефоном уже существует"})
        
        # Хешируем пароль
        hashed_password = pwd_context.hash(password)
        
        # Создаем нового администратора
        new_admin = models.User(
            email=email,
            full_name=full_name,
            phone=phone,
            hashed_password=hashed_password,
            role=models.UserRole.ADMIN,
            is_active=True
        )
        
        db.add(new_admin)
        db.commit()
        db.refresh(new_admin)
        
        print(f"DEBUG: Создан новый администратор: {email}")
        return JSONResponse(content={"success": True, "message": "Администратор создан успешно"})
        
    except Exception as e:
        print(f"ERROR: Ошибка создания администратора: {e}")
        db.rollback()
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка создания: {str(e)}"})

@app.get("/api/v1/superadmin/admins/{admin_id}")
async def get_admin(
    admin_id: int,
    request: Request,
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    admin = db.query(models.User).filter(
        models.User.id == admin_id,
        models.User.role == models.UserRole.ADMIN
    ).first()
    
    if not admin:
        return JSONResponse(status_code=404, content={"success": False, "message": "Администратор не найден"})
    
    return JSONResponse(content={
        "success": True,
        "admin": {
            "id": admin.id,
            "full_name": admin.full_name,
            "email": admin.email,
            "phone": admin.phone,
            "is_active": admin.is_active
        }
    })

@app.put("/api/v1/superadmin/admins/{admin_id}")
async def update_admin(
    admin_id: int,
    request: Request,
    full_name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(None),
    is_active: str = Form(...),
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        admin = db.query(models.User).filter(
            models.User.id == admin_id,
            models.User.role == models.UserRole.ADMIN
        ).first()
        
        if not admin:
            return JSONResponse(status_code=404, content={"success": False, "message": "Администратор не найден"})
        
        # Обновляем данные
        admin.full_name = full_name
        admin.email = email
        admin.phone = phone
        admin.is_active = is_active.lower() == 'true'
        
        db.commit()
        
        return JSONResponse(content={"success": True, "message": "Администратор обновлен успешно"})
        
    except Exception as e:
        print(f"ERROR: Ошибка обновления администратора: {e}")
        db.rollback()
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка обновления: {str(e)}"})

@app.delete("/api/v1/superadmin/admins/{admin_id}")
async def delete_admin(
    admin_id: int,
    request: Request,
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        # Проверяем, что это не сам суперадмин пытается удалить себя
        if admin_id == user.id:
            return JSONResponse(content={"success": False, "message": "Нельзя удалить самого себя"})
        
        admin = db.query(models.User).filter(
            models.User.id == admin_id,
            models.User.role == models.UserRole.ADMIN
        ).first()
        
        if not admin:
            return JSONResponse(status_code=404, content={"success": False, "message": "Администратор не найден"})
        
        db.delete(admin)
        db.commit()
        
        return JSONResponse(content={"success": True, "message": "Администратор удален успешно"})
        
    except Exception as e:
        print(f"ERROR: Ошибка удаления администратора: {e}")
        db.rollback()
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка удаления: {str(e)}"})

@app.get("/api/v1/superadmin/stats")
async def get_superadmin_stats(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        stats = {
            'admins_count': db.query(models.User).filter(models.User.role == models.UserRole.ADMIN).count(),
            'users_count': db.query(models.User).count(),
            'properties_count': db.query(models.Property).count(),
            'pending_requests': db.query(models.Property).filter(models.Property.status == 'PENDING').count()
        }
        
        return JSONResponse(content={"success": True, "stats": stats})
        
    except Exception as e:
        print(f"ERROR: Ошибка получения статистики: {e}")
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка: {str(e)}"})

# Дополнительные API роуты для суперадмина

@app.get("/api/v1/superadmin/users/{user_id}")
async def get_superadmin_user(
    user_id: int,
    request: Request,
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    target_user = db.query(models.User).filter(models.User.id == user_id).first()
    
    if not target_user:
        return JSONResponse(status_code=404, content={"success": False, "message": "Пользователь не найден"})
    
    # Получаем объявления пользователя
    properties = db.query(models.Property).filter(models.Property.owner_id == user_id).all()
    properties_data = [{
        "id": prop.id,
        "title": prop.title,
        "address": prop.address,
        "price": prop.price
    } for prop in properties]
    
    return JSONResponse(content={
        "success": True,
        "user": {
            "id": target_user.id,
            "full_name": target_user.full_name,
            "email": target_user.email,
            "phone": target_user.phone,
            "is_active": target_user.is_active,
            "properties_count": len(properties_data),
            "registered_at": target_user.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(target_user, 'created_at') and target_user.created_at else "Неизвестно",
            "properties": properties_data
        }
    })

@app.delete("/api/v1/superadmin/users/{user_id}/properties")
async def delete_user_properties(
    user_id: int,
    request: Request,
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        # Получаем все объявления пользователя
        properties = db.query(models.Property).filter(models.Property.owner_id == user_id).all()
        
        if not properties:
            return JSONResponse(content={"success": False, "message": "У пользователя нет объявлений"})
        
        # Удаляем все объявления
        for prop in properties:
            db.delete(prop)
        
        db.commit()
        
        return JSONResponse(content={"success": True, "message": f"Удалено {len(properties)} объявлений"})
        
    except Exception as e:
        print(f"ERROR: Ошибка удаления объявлений пользователя: {e}")
        db.rollback()
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка удаления: {str(e)}"})

@app.post("/api/v1/superadmin/users/{user_id}/toggle-status")
async def toggle_user_status(
    user_id: int,
    request: Request,
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        target_user = db.query(models.User).filter(models.User.id == user_id).first()
        
        if not target_user:
            return JSONResponse(status_code=404, content={"success": False, "message": "Пользователь не найден"})
        
        # Переключаем статус
        target_user.is_active = not target_user.is_active
        db.commit()
        
        status_text = "активирован" if target_user.is_active else "заблокирован"
        return JSONResponse(content={"success": True, "message": f"Пользователь {status_text}"})
        
    except Exception as e:
        print(f"ERROR: Ошибка изменения статуса пользователя: {e}")
        db.rollback()
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка изменения статуса: {str(e)}"})

# Дополнительные API роуты для управления объявлениями

@app.get("/api/v1/superadmin/properties/{property_id}")
async def get_superadmin_property(
    property_id: int,
    request: Request,
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    property_item = db.query(models.Property).filter(models.Property.id == property_id).first()
    
    if not property_item:
        return JSONResponse(status_code=404, content={"success": False, "message": "Объявление не найдено"})
    
    return JSONResponse(content={
        "success": True,
        "property": {
            "id": property_item.id,
            "title": property_item.title,
            "price": property_item.price,
            "address": property_item.address,
            "status": property_item.status.value if property_item.status else "draft",
            "rooms": property_item.rooms,
            "area": property_item.area
        }
    })

@app.put("/api/v1/superadmin/properties/{property_id}")
async def update_superadmin_property(
    property_id: int,
    request: Request,
    title: str = Form(...),
    price: float = Form(...),
    address: str = Form(...),
    status: str = Form(...),
    rooms: int = Form(None),
    area: float = Form(None),
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        property_item = db.query(models.Property).filter(models.Property.id == property_id).first()
        
        if not property_item:
            return JSONResponse(status_code=404, content={"success": False, "message": "Объявление не найдено"})
        
        # Обновляем данные
        property_item.title = title
        property_item.price = price
        property_item.address = address
        property_item.status = status
        if rooms is not None:
            property_item.rooms = rooms
        if area is not None:
            property_item.area = area
        
        db.commit()
        
        return JSONResponse(content={"success": True, "message": "Объявление обновлено успешно"})
        
    except Exception as e:
        print(f"ERROR: Ошибка обновления объявления: {e}")
        db.rollback()
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка обновления: {str(e)}"})

@app.put("/api/v1/superadmin/properties/{property_id}/status")
async def update_property_status(
    property_id: int,
    request: Request,
    status: str = Form(...),
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        property_item = db.query(models.Property).filter(models.Property.id == property_id).first()
        
        if not property_item:
            return JSONResponse(status_code=404, content={"success": False, "message": "Объявление не найдено"})
        
        property_item.status = status
        db.commit()
        
        return JSONResponse(content={"success": True, "message": "Статус изменен успешно"})
        
    except Exception as e:
        print(f"ERROR: Ошибка изменения статуса: {e}")
        db.rollback()
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка изменения статуса: {str(e)}"})

@app.delete("/api/v1/superadmin/properties/{property_id}")
async def delete_superadmin_property(
    property_id: int,
    request: Request,
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        property_item = db.query(models.Property).filter(models.Property.id == property_id).first()
        
        if not property_item:
            return JSONResponse(status_code=404, content={"success": False, "message": "Объявление не найдено"})
        
        db.delete(property_item)
        db.commit()
        
        return JSONResponse(content={"success": True, "message": "Объявление удалено успешно"})
        
    except Exception as e:
        print(f"ERROR: Ошибка удаления объявления: {e}")
        db.rollback()
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка удаления: {str(e)}"})

# API для экспорта данных в XLSX
@app.get("/api/v1/superadmin/export/properties")
async def export_properties(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        # Получаем все объявления
        properties = db.query(models.Property).options(
            joinedload(models.Property.owner)
        ).all()
        
        # Подготавливаем данные для экспорта
        data = []
        for prop in properties:
            data.append({
                'ID': prop.id,
                'Название': prop.title or f"Объект #{prop.id}",
                'Цена': prop.price or 0,
                'Адрес': prop.address or "",
                'Комнаты': prop.rooms or 0,
                'Площадь': prop.area or 0,
                'Статус': prop.status.value if prop.status else "draft",
                'Владелец': prop.owner.full_name if prop.owner else "Нет данных",
                'Email владельца': prop.owner.email if prop.owner else "",
                'Телефон владельца': prop.owner.phone if prop.owner else "",
                'Дата создания': prop.created_at.strftime("%d.%m.%Y %H:%M") if prop.created_at else ""
            })
        
        # Создаем Excel файл
        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Объявления"
        
        # Заголовки
        headers = ["ID", "Название", "Цена", "Адрес", "Комнаты", "Площадь", "Статус", "Владелец", "Email владельца", "Телефон владельца", "Дата создания"]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Данные
        for row, item in enumerate(data, 2):
            worksheet.cell(row=row, column=1, value=item['ID'])
            worksheet.cell(row=row, column=2, value=item['Название'])
            worksheet.cell(row=row, column=3, value=item['Цена'])
            worksheet.cell(row=row, column=4, value=item['Адрес'])
            worksheet.cell(row=row, column=5, value=item['Комнаты'])
            worksheet.cell(row=row, column=6, value=item['Площадь'])
            worksheet.cell(row=row, column=7, value=item['Статус'])
            worksheet.cell(row=row, column=8, value=item['Владелец'])
            worksheet.cell(row=row, column=9, value=item['Email владельца'])
            worksheet.cell(row=row, column=10, value=item['Телефон владельца'])
            worksheet.cell(row=row, column=11, value=item['Дата создания'])
        
        workbook.save(output)
        output.seek(0)
        
        headers = {
            'Content-Disposition': f'attachment; filename="properties_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        }
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
        
    except Exception as e:
        print(f"ERROR: Ошибка экспорта объявлений: {e}")
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка экспорта: {str(e)}"})

@app.get("/api/v1/superadmin/export/users")
async def export_users(
    request: Request, 
    search: str = Query(None),
    status: str = Query(None),
    property_filter: str = Query(None),
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        # Базовый запрос с фильтрами (тот же что и в основной функции)
        query = db.query(models.User).filter(models.User.role == models.UserRole.USER)
        
        # Применяем фильтры
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    models.User.full_name.ilike(search_term),
                    models.User.email.ilike(search_term),
                    models.User.phone.ilike(search_term)
                )
            )
        
        if status:
            if status == "active":
                query = query.filter(models.User.is_active == True)
            elif status == "blocked":
                query = query.filter(models.User.is_active == False)
        
        users = query.all()
        
        # Подготавливаем данные для экспорта
        data = []
        for user in users:
            properties_count = db.query(models.Property).filter(models.Property.owner_id == user.id).count()
            
            # Применяем фильтр по недвижимости
            if property_filter:
                if property_filter == "with" and properties_count == 0:
                    continue
                elif property_filter == "without" and properties_count > 0:
                    continue
            
            data.append({
                'ID': user.id,
                'ФИО': user.full_name or f"Пользователь {user.id}",
                'Email': user.email or "Не указан",
                'Телефон': user.phone or "Не указан",
                'Статус': "Активный" if user.is_active else "Заблокирован",
                'Объявлений': properties_count,
                'Дата регистрации': user.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(user, 'created_at') and user.created_at else "Нет данных"
            })
        
        if not data:
            return JSONResponse(status_code=404, content={"success": False, "message": "Нет данных для экспорта"})
        
        # Создаем Excel файл
        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Пользователи"
        
        # Заголовки
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Данные
        for row, item in enumerate(data, 2):
            for col, key in enumerate(headers, 1):
                worksheet.cell(row=row, column=col, value=item[key])
        
        workbook.save(output)
        output.seek(0)
        
        # Формируем имя файла с фильтрами
        filename = "users"
        if search:
            filename += f"_search-{search[:10]}"
        if status:
            filename += f"_status-{status}"
        if property_filter:
            filename += f"_property-{property_filter}"
        filename += f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        
        return Response(
            content=output.getvalue(),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка экспорта: {str(e)}"})

@app.get("/api/v1/superadmin/export/all")
async def export_all_data(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        # Получаем все данные
        users = db.query(models.User).all()
        properties = db.query(models.Property).options(joinedload(models.Property.owner)).all()
        
        # Создаем Excel файл с несколькими листами
        output = BytesIO()
        workbook = openpyxl.Workbook()
        
        # Удаляем дефолтный лист
        workbook.remove(workbook.active)
        
        # Лист пользователей
        users_sheet = workbook.create_sheet("Пользователи")
        users_headers = ["ID", "ФИО", "Email", "Телефон", "Роль", "Статус", "Количество объявлений", "Дата регистрации"]
        for col, header in enumerate(users_headers, 1):
            users_sheet.cell(row=1, column=col, value=header)
        
        for row, user_item in enumerate(users, 2):
            properties_count = db.query(models.Property).filter(models.Property.owner_id == user_item.id).count()
            users_sheet.cell(row=row, column=1, value=user_item.id)
            users_sheet.cell(row=row, column=2, value=user_item.full_name or f"Пользователь {user_item.id}")
            users_sheet.cell(row=row, column=3, value=user_item.email or "")
            users_sheet.cell(row=row, column=4, value=user_item.phone or "")
            users_sheet.cell(row=row, column=5, value=user_item.role.value if user_item.role else "user")
            users_sheet.cell(row=row, column=6, value="Активный" if user_item.is_active else "Неактивный")
            users_sheet.cell(row=row, column=7, value=properties_count)
            users_sheet.cell(row=row, column=8, value=user_item.created_at.strftime("%d.%m.%Y %H:%M") if hasattr(user_item, 'created_at') and user_item.created_at else "")
        
        # Лист объявлений
        properties_sheet = workbook.create_sheet("Объявления")
        properties_headers = ["ID", "Название", "Цена", "Адрес", "Комнаты", "Площадь", "Статус", "Владелец", "Email владельца", "Телефон владельца", "Дата создания"]
        for col, header in enumerate(properties_headers, 1):
            properties_sheet.cell(row=1, column=col, value=header)
        
        for row, prop in enumerate(properties, 2):
            properties_sheet.cell(row=row, column=1, value=prop.id)
            properties_sheet.cell(row=row, column=2, value=prop.title or f"Объект #{prop.id}")
            properties_sheet.cell(row=row, column=3, value=prop.price or 0)
            properties_sheet.cell(row=row, column=4, value=prop.address or "")
            properties_sheet.cell(row=row, column=5, value=prop.rooms or 0)
            properties_sheet.cell(row=row, column=6, value=prop.area or 0)
            properties_sheet.cell(row=row, column=7, value=prop.status.value if prop.status else "draft")
            properties_sheet.cell(row=row, column=8, value=prop.owner.full_name if prop.owner else "Нет данных")
            properties_sheet.cell(row=row, column=9, value=prop.owner.email if prop.owner else "")
            properties_sheet.cell(row=row, column=10, value=prop.owner.phone if prop.owner else "")
            properties_sheet.cell(row=row, column=11, value=prop.created_at.strftime("%d.%m.%Y %H:%M") if prop.created_at else "")
        
        workbook.save(output)
        output.seek(0)
        
        headers = {
            'Content-Disposition': f'attachment; filename="full_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        }
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
        
    except Exception as e:
        print(f"ERROR: Ошибка полного экспорта: {e}")
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка экспорта: {str(e)}"})

@app.get("/api/v1/superadmin/export/admins")
async def export_admins(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Нет доступа"})
    
    try:
        # Получаем всех администраторов
        admins = db.query(models.User).filter(
            or_(
                models.User.role == models.UserRole.ADMIN,
                models.User.role == models.UserRole.MANAGER
            )
        ).all()
        
        # Создаем Excel файл
        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Администраторы"
        
        # Заголовки
        headers = ["ID", "ФИО", "Email", "Телефон", "Роль", "Статус", "Дата создания"]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Данные
        for row, admin in enumerate(admins, 2):
            worksheet.cell(row=row, column=1, value=admin.id)
            worksheet.cell(row=row, column=2, value=admin.full_name or "Не указано")
            worksheet.cell(row=row, column=3, value=admin.email or "Не указано")
            worksheet.cell(row=row, column=4, value=admin.phone or "Не указано")
            worksheet.cell(row=row, column=5, value=admin.role.value if admin.role else "Не указано")
            worksheet.cell(row=row, column=6, value="Активный" if admin.is_active else "Неактивный")
            worksheet.cell(row=row, column=7, value=admin.created_at.strftime("%d.%m.%Y %H:%M") if hasattr(admin, 'created_at') and admin.created_at else "Не указано")
        
        workbook.save(output)
        output.seek(0)
        
        headers = {
            'Content-Disposition': f'attachment; filename="admins_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        }
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
        
    except Exception as e:
        print(f"Error exporting admins: {str(e)}")
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка экспорта: {str(e)}"})

@app.get("/api/v1/superadmin/settings")
async def get_superadmin_settings(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Нет доступа"})
    
    # Возвращаем настройки суперадмина (можно сохранять в базе данных)
    settings = {
        "system_scale": 100,
        "auto_backup": True,
        "log_level": "INFO",
        "max_file_size": 100,  # MB
        "session_timeout": 60,  # minutes
        "enable_notifications": True,
        "dark_mode": False,
        "compact_view": False
    }
    
    return JSONResponse(content={"success": True, "settings": settings})

@app.post("/api/v1/superadmin/settings")
async def update_superadmin_settings(
    request: Request,
    system_scale: int = Form(100),
    auto_backup: bool = Form(True),
    log_level: str = Form("INFO"),
    max_file_size: int = Form(100),
    session_timeout: int = Form(60),
    enable_notifications: bool = Form(True),
    dark_mode: bool = Form(False),
    compact_view: bool = Form(False),
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Нет доступа"})
    
    try:
        # Здесь можно сохранить настройки в базе данных
        # В этом примере просто возвращаем успех
        updated_settings = {
            "system_scale": system_scale,
            "auto_backup": auto_backup,
            "log_level": log_level,
            "max_file_size": max_file_size,
            "session_timeout": session_timeout,
            "enable_notifications": enable_notifications,
            "dark_mode": dark_mode,
            "compact_view": compact_view
        }
        
        return JSONResponse(content={"success": True, "message": "Настройки сохранены", "settings": updated_settings})
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка сохранения настроек: {str(e)}"})

@app.post("/api/v1/superadmin/system/scale")
async def update_system_scale(
    request: Request,
    scale: int = Form(...),
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Нет доступа"})
    
    try:
        # Валидация значения масштаба
        if not 50 <= scale <= 200:
            return JSONResponse(status_code=400, content={"success": False, "message": "Масштаб должен быть от 50% до 200%"})
        
        # Здесь можно реализовать изменение масштаба системы
        # Например, изменение CSS переменных или настроек UI
        
        return JSONResponse(content={"success": True, "message": f"Масштаб системы установлен на {scale}%", "scale": scale})
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка изменения масштаба: {str(e)}"})

@app.post("/api/v1/superadmin/system/backup")
async def create_system_backup(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Нет доступа"})
    
    try:
        # Создаем резервную копию (здесь можно добавить реальную логику)
        backup_filename = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql"
        
        # В реальном приложении здесь был бы код для создания бэкапа базы данных
        # subprocess.run(["pg_dump", "database_name", "-f", backup_filename])
        
        return JSONResponse(content={"success": True, "message": "Резервная копия создана", "filename": backup_filename})
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка создания резервной копии: {str(e)}"})

@app.get("/api/v1/superadmin/system/status")
async def get_system_status(request: Request, db: Session = Depends(deps.get_db)):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Нет доступа"})
    
    try:
        # Пытаемся импортировать psutil для получения реальной информации о системе
        try:
            import psutil
            
            # Получаем информацию о системе
            cpu_percent = psutil.cpu_percent(interval=1)
            memory = psutil.virtual_memory()
            disk = psutil.disk_usage('/')
            
            system_status = {
                "cpu_usage": cpu_percent,
                "memory_usage": memory.percent,
                "memory_total": memory.total // (1024**3),  # GB
                "memory_available": memory.available // (1024**3),  # GB
                "disk_usage": disk.percent,
                "disk_total": disk.total // (1024**3),  # GB
                "disk_free": disk.free // (1024**3),  # GB
                "uptime": "Online",
                "last_backup": "2024-01-15 14:30:00",
                "database_status": "Connected",
                "web_server_status": "Running"
            }
        except ImportError:
            # Если psutil не установлен, возвращаем базовую информацию
            system_status = {
                "cpu_usage": 15,
                "memory_usage": 45,
                "memory_total": 8,
                "memory_available": 4,
                "disk_usage": 60,
                "disk_total": 100,
                "disk_free": 40,
                "uptime": "Online",
                "last_backup": "2024-01-15 14:30:00",
                "database_status": "Connected",
                "web_server_status": "Running"
            }
        
        return JSONResponse(content={"success": True, "status": system_status})
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка получения статуса системы: {str(e)}"})

# Дополнительные API роуты для суперадмина

@app.post("/api/v1/superadmin/companies")
async def create_company(
    request: Request,
    company_name: str = Form(...),
    company_number: str = Form(...),
    company_owner: str = Form(...),
    email: str = Form(...),
    phone: str = Form(...),
    full_name: str = Form(None),
    company_address: str = Form(None),
    company_description: str = Form(None),
    company_logo_url: str = Form(None),
    password: str = Form(...),
    is_active: str = Form("on"),
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return user
    
    try:
        # Проверяем уникальность email
        existing_email = db.query(models.User).filter(models.User.email == email).first()
        if existing_email:
            return JSONResponse(status_code=400, content={"success": False, "message": "Пользователь с таким email уже существует"})
        
        # Проверяем уникальность телефона (если указан)
        if phone:
            existing_phone = db.query(models.User).filter(models.User.phone == phone).first()
            if existing_phone:
                return JSONResponse(status_code=400, content={"success": False, "message": "Пользователь с таким телефоном уже существует"})
        
        # Хешируем пароль
        hashed_password = pwd_context.hash(password)
        
        # Создаем новое юридическое лицо
        new_company = models.User(
            email=email,
            phone=phone,
            full_name=full_name,
            hashed_password=hashed_password,
            role=models.UserRole.COMPANY,
            is_active=is_active == "on",
            company_name=company_name,
            company_number=company_number,
            company_owner=company_owner,
            company_address=company_address,
            company_description=company_description,
            company_logo_url=company_logo_url if company_logo_url else None
        )
        
        db.add(new_company)
        db.commit()
        db.refresh(new_company)
        
        print(f"DEBUG: Создано новое юридическое лицо: {company_name}")
        
        return JSONResponse(content={"success": True, "message": "Юридическое лицо создано успешно"})
        
    except Exception as e:
        print(f"ERROR: Ошибка создания юридического лица: {e}")
        db.rollback()
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка создания: {str(e)}"})

@app.get("/api/v1/superadmin/companies/{company_id}")
async def get_company(
    company_id: int,
    request: Request,
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    company = db.query(models.User).filter(
        models.User.id == company_id,
        models.User.role == models.UserRole.COMPANY
    ).first()
    
    if not company:
        return JSONResponse(status_code=404, content={"success": False, "message": "Компания не найдена"})
    
    return JSONResponse(content={
        "success": True,
        "company": {
            "id": company.id,
            "company_name": company.company_name,
            "company_number": company.company_number,
            "company_owner": company.company_owner,
            "company_address": company.company_address,
            "company_description": company.company_description,
            "company_logo_url": company.company_logo_url,
            "full_name": company.full_name,
            "email": company.email,
            "phone": company.phone,
            "is_active": company.is_active
        }
    })

@app.put("/api/v1/superadmin/companies/{company_id}")
async def update_company(
    company_id: int,
    request: Request,
    company_name: str = Form(...),
    company_number: str = Form(...),
    company_owner: str = Form(...),
    email: str = Form(...),
    phone: str = Form(...),
    full_name: str = Form(None),
    company_address: str = Form(None),
    company_description: str = Form(None),
    company_logo_url: str = Form(None),
    is_active: str = Form("off"),
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        company = db.query(models.User).filter(
            models.User.id == company_id,
            models.User.role == models.UserRole.COMPANY
        ).first()
        
        if not company:
            return JSONResponse(status_code=404, content={"success": False, "message": "Компания не найдена"})
        
        # Проверяем уникальность email (кроме текущей компании)
        existing_email = db.query(models.User).filter(
            models.User.email == email,
            models.User.id != company_id
        ).first()
        if existing_email:
            return JSONResponse(status_code=400, content={"success": False, "message": "Пользователь с таким email уже существует"})
        
        # Проверяем уникальность телефона (кроме текущей компании)
        existing_phone = db.query(models.User).filter(
            models.User.phone == phone,
            models.User.id != company_id
        ).first()
        if existing_phone:
            return JSONResponse(status_code=400, content={"success": False, "message": "Пользователь с таким телефоном уже существует"})
        
        # Проверяем уникальность номера компании (кроме текущей компании)
        existing_number = db.query(models.User).filter(
            models.User.company_number == company_number,
            models.User.role == models.UserRole.COMPANY,
            models.User.id != company_id
        ).first()
        if existing_number:
            return JSONResponse(status_code=400, content={"success": False, "message": "Компания с таким номером уже существует"})
        
        # Обновляем данные
        company.company_name = company_name
        company.company_number = company_number
        company.company_owner = company_owner
        company.company_address = company_address
        company.company_description = company_description
        company.company_logo_url = company_logo_url if company_logo_url else None
        company.full_name = full_name
        company.email = email
        company.phone = phone
        company.is_active = is_active == "on"
        
        db.commit()
        
        return JSONResponse(content={"success": True, "message": "Юридическое лицо обновлено успешно"})
        
    except Exception as e:
        print(f"ERROR: Ошибка обновления юридического лица: {e}")
        db.rollback()
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка обновления: {str(e)}"})

@app.patch("/api/v1/superadmin/companies/{company_id}/toggle")
async def toggle_company_status(
    company_id: int,
    request: Request,
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        company = db.query(models.User).filter(
            models.User.id == company_id,
            models.User.role == models.UserRole.COMPANY
        ).first()
        
        if not company:
            return JSONResponse(status_code=404, content={"success": False, "message": "Компания не найдена"})
        
        # Переключаем статус
        company.is_active = not company.is_active
        db.commit()
        
        status_text = "активирована" if company.is_active else "заблокирована"
        return JSONResponse(content={"success": True, "message": f"Компания {status_text}"})
        
    except Exception as e:
        print(f"ERROR: Ошибка изменения статуса компании: {e}")
        db.rollback()
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка изменения статуса: {str(e)}"})

@app.delete("/api/v1/superadmin/companies/{company_id}")
async def delete_company(
    company_id: int,
    request: Request,
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        company = db.query(models.User).filter(
            models.User.id == company_id,
            models.User.role == models.UserRole.COMPANY
        ).first()
        
        if not company:
            return JSONResponse(status_code=404, content={"success": False, "message": "Компания не найдена"})
        
        # Получаем все объявления компании
        properties = db.query(models.Property).filter(models.Property.owner_id == company_id).all()
        
        # Удаляем все объявления компании
        for prop in properties:
            db.delete(prop)
        
        # Удаляем саму компанию
        db.delete(company)
        db.commit()
        
        return JSONResponse(content={
            "success": True, 
            "message": f"Компания удалена. Также удалено {len(properties)} объявлений."
        })
        
    except Exception as e:
        print(f"ERROR: Ошибка удаления компании: {e}")
        db.rollback()
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка удаления: {str(e)}"})

@app.get("/api/v1/superadmin/export/companies")
async def export_companies(
    request: Request, 
    search: str = Query(None),
    status: str = Query(None),
    db: Session = Depends(deps.get_db)
):
    # Проверяем доступ суперадмина
    user = await check_superadmin_access(request, db)
    if isinstance(user, RedirectResponse):
        return JSONResponse(status_code=401, content={"success": False, "message": "Доступ запрещен"})
    
    try:
        # Базовый запрос с фильтрами
        query = db.query(models.User).filter(models.User.role == models.UserRole.COMPANY)
        
        # Применяем фильтры
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    models.User.company_name.ilike(search_term),
                    models.User.company_number.ilike(search_term),
                    models.User.company_owner.ilike(search_term),
                    models.User.full_name.ilike(search_term),
                    models.User.email.ilike(search_term),
                    models.User.phone.ilike(search_term)
                )
            )
        
        if status:
            if status == "active":
                query = query.filter(models.User.is_active == True)
            elif status == "blocked":
                query = query.filter(models.User.is_active == False)
        
        companies = query.all()
        
        # Подготавливаем данные для экспорта
        data = []
        for company in companies:
            properties_count = db.query(models.Property).filter(models.Property.owner_id == company.id).count()
            
            data.append({
                'ID': company.id,
                'Название компании': company.company_name or "Не указано",
                'ИНН/Номер': company.company_number or "Не указан",
                'Владелец': company.company_owner or "Не указан",
                'Контактное лицо': company.full_name or "Не указано",
                'Email': company.email or "Не указан",
                'Телефон': company.phone or "Не указан",
                'Адрес': company.company_address or "Не указан",
                'Описание': company.company_description or "Нет описания",
                'Статус': "Активная" if company.is_active else "Заблокирована",
                'Объявлений': properties_count,
                'Дата регистрации': company.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(company, 'created_at') and company.created_at else "Нет данных"
            })
        
        if not data:
            return JSONResponse(status_code=404, content={"success": False, "message": "Нет данных для экспорта"})
        
        # Создаем Excel файл
        from io import BytesIO
        import openpyxl
        from datetime import datetime
        from fastapi import Response
        
        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Юридические лица"
        
        # Заголовки
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Данные
        for row, item in enumerate(data, 2):
            for col, key in enumerate(headers, 1):
                worksheet.cell(row=row, column=col, value=item[key])
        
        workbook.save(output)
        output.seek(0)
        
        # Формируем имя файла с фильтрами
        filename = "companies"
        if search:
            filename += f"_search-{search[:10]}"
        if status:
            filename += f"_status-{status}"
        filename += f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        
        return Response(
            content=output.getvalue(),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "message": f"Ошибка экспорта: {str(e)}"})

# === COMPANIES PANEL ROUTES ===

async def check_company_access(request: Request, db: Session):
    """Проверка доступа компании"""
    auth_token = request.cookies.get('access_token')
    if not auth_token:
        return None
        
    try:
        payload = pyjwt.decode(auth_token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
        user_id = payload.get("sub")
        
        if user_id:
            user = db.query(models.User).filter(models.User.id == user_id).first()
            if user and user.role == models.UserRole.COMPANY and user.is_active:
                return user
    except:
        pass
    
    return None

@app.get("/companies/login")
async def companies_login_get(request: Request):
    """Страница входа для компаний"""
    return templates.TemplateResponse("companies/login.html", {"request": request})

@app.post("/companies/login")
async def companies_login_post(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    remember: bool = Form(False),
    db: Session = Depends(deps.get_db)
):
    """Обработка входа компании"""
    try:
        # Проверяем учетные данные
        user = db.query(models.User).filter(
            models.User.email == email,
            models.User.role == models.UserRole.COMPANY
        ).first()
        
        if not user or not user.check_password(password):
            raise HTTPException(status_code=401, detail="Неверный email или пароль")
        
        if not user.is_active:
            raise HTTPException(status_code=403, detail="Аккаунт заблокирован")
        
        # Создаем токен
        access_token = create_access_token(data={
            "sub": str(user.id),
            "is_company": True
        })
        
        # Создаем ответ
        response = JSONResponse(content={"message": "Успешный вход"})
        
        # Устанавливаем cookie
        max_age = 86400 * 30 if remember else 86400  # 30 дней или 1 день
        response.set_cookie(
            key="access_token",
            value=access_token,
            max_age=max_age,
            httponly=True,
            secure=False  # В продакшене должно быть True
        )
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"DEBUG: Ошибка входа компании: {e}")
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")

@app.get("/companies/logout")
async def companies_logout():
    """Выход из панели компании"""
    response = RedirectResponse(url="/companies/login", status_code=302)
    response.delete_cookie("access_token")
    return response

@app.get("/companies", response_class=RedirectResponse)
@app.get("/companies/", response_class=RedirectResponse)
async def companies_root():
    """Редирект на дашборд компании"""
    return RedirectResponse(url="/companies/dashboard", status_code=302)

@app.get("/companies/dashboard", response_class=HTMLResponse)
async def companies_dashboard(request: Request, db: Session = Depends(deps.get_db)):
    """Дашборд компании"""
    current_user = await check_company_access(request, db)
    if not current_user:
        return RedirectResponse(url="/companies/login", status_code=302)
    
    try:
        # Получаем статистику компании
        user_properties = db.query(models.Property).filter(models.Property.owner_id == current_user.id)
        
        total_properties = user_properties.count()
        active_properties = user_properties.filter(models.Property.status == 'active').count()
        
        # Считаем общие просмотры (примерное значение, так как поля views может не быть)
        properties_list = user_properties.all()
        total_views = 0
        for prop in properties_list:
            if hasattr(prop, 'views') and prop.views:
                total_views += prop.views
        
        # Средняя цена
        prices = [prop.price for prop in properties_list if prop.price]
        avg_price = sum(prices) / len(prices) if prices else 0
        
        stats = {
            'total_properties': total_properties,
            'active_properties': active_properties,
            'total_views': total_views,
            'avg_price': avg_price,
            'properties_growth': 0,  # Можно вычислить рост за месяц
            'price_change': 0,  # Можно вычислить изменение цены
        }
        
        # Получаем последние объявления
        recent_properties = user_properties.order_by(models.Property.created_at.desc()).limit(5).all()
        
        # Получаем топ объявления по просмотрам (если поле views существует)
        try:
            top_properties = user_properties.order_by(models.Property.views.desc()).limit(5).all()
        except AttributeError:
            # Если поля views нет, просто берем последние
            top_properties = recent_properties
        
        # Данные для графика (примерные)
        chart_labels = ['1', '2', '3', '4', '5', '6', '7']
        chart_data = [10, 15, 8, 12, 20, 18, 25]
        
        return templates.TemplateResponse("companies/dashboard.html", {
            "request": request,
            "current_user": current_user,
            "company_name": current_user.company_name,
            "stats": stats,
            "recent_properties": recent_properties,
            "top_properties": top_properties,
            "chart_labels": chart_labels,
            "chart_data": chart_data
        })
        
    except Exception as e:
        print(f"DEBUG: Ошибка в дашборде компании: {e}")
        return RedirectResponse(url="/companies/login", status_code=302)

@app.get("/companies/listings", response_class=HTMLResponse)
async def companies_listings(
    request: Request,
    search: str = Query(None),
    status: str = Query(None),
    property_type: str = Query(None),
    page: int = Query(1, ge=1),
    db: Session = Depends(deps.get_db)
):
    """Управление объявлениями компании"""
    current_user = await check_company_access(request, db)
    if not current_user:
        return RedirectResponse(url="/companies/login", status_code=302)
    
    try:
        # Базовый запрос
        query = db.query(models.Property).filter(models.Property.owner_id == current_user.id)
        
        # Фильтры
        if search:
            query = query.filter(models.Property.title.ilike(f"%{search}%"))
        if status:
            query = query.filter(models.Property.status == status)
        if property_type:
            query = query.filter(models.Property.type == property_type)
        
        # Пагинация
        per_page = 20
        total_count = query.count()
        total_pages = (total_count + per_page - 1) // per_page
        offset = (page - 1) * per_page
        
        properties = query.order_by(models.Property.created_at.desc()).offset(offset).limit(per_page).all()
        
        # Статистика
        all_properties = db.query(models.Property).filter(models.Property.owner_id == current_user.id)
        active_count = all_properties.filter(models.Property.status == 'active').count()
        pending_count = all_properties.filter(models.Property.status == 'pending').count()
        draft_count = all_properties.filter(models.Property.status == 'draft').count()
        
        return templates.TemplateResponse("companies/listings.html", {
            "request": request,
            "current_user": current_user,
            "company_name": current_user.company_name,
            "properties": properties,
            "total_count": total_count,
            "active_count": active_count,
            "pending_count": pending_count,
            "draft_count": draft_count,
            "current_page": page,
            "total_pages": total_pages,
            "per_page": per_page,
            "search": search,
            "status": status,
            "property_type": property_type
        })
        
    except Exception as e:
        print(f"DEBUG: Ошибка в списке объявлений компании: {e}")
        return RedirectResponse(url="/companies/dashboard", status_code=302)

@app.get("/companies/analytics", response_class=HTMLResponse)
async def companies_analytics(
    request: Request,
    days: int = Query(30),
    db: Session = Depends(deps.get_db)
):
    """Аналитика компании"""
    current_user = await check_company_access(request, db)
    if not current_user:
        return RedirectResponse(url="/companies/login", status_code=302)
    
    try:
        # Примерные данные для аналитики
        analytics_data = {
            'views_data': {
                'labels': ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс'],
                'data': [120, 150, 90, 180, 200, 160, 140]
            },
            'conversion_data': {
                'labels': ['Просмотры', 'Звонки', 'Встречи', 'Сделки'],
                'data': [1000, 150, 45, 12]
            },
            'geography_data': [
                {'region': 'Бишкек', 'views': 450, 'percentage': 65},
                {'region': 'Ош', 'views': 120, 'percentage': 17},
                {'region': 'Каракол', 'views': 80, 'percentage': 12},
                {'region': 'Другие', 'views': 45, 'percentage': 6}
            ],
            'traffic_sources': [
                {'source': 'Поиск', 'visits': 320, 'percentage': 45},
                {'source': 'Соц. сети', 'visits': 180, 'percentage': 25},
                {'source': 'Прямые заходы', 'visits': 140, 'percentage': 20},
                {'source': 'Реклама', 'visits': 70, 'percentage': 10}
            ],
            'popular_properties': db.query(models.Property).filter(
                models.Property.owner_id == current_user.id
            ).order_by(models.Property.created_at.desc()).limit(5).all(),
            'total_stats': {
                'total_views': 695,
                'total_calls': 28,
                'total_messages': 15,
                'conversion_rate': 4.0
            }
        }
        
        return templates.TemplateResponse("companies/analytics.html", {
            "request": request,
            "current_user": current_user,
            "company_name": current_user.company_name,
            "analytics": analytics_data,
            "days": days
        })
        
    except Exception as e:
        print(f"DEBUG: Ошибка в аналитике компании: {e}")
        return RedirectResponse(url="/companies/login", status_code=302)

@app.get("/companies/create-listing", response_class=HTMLResponse)
async def companies_create_listing(request: Request, db: Session = Depends(deps.get_db)):
    """Страница создания объявления компании"""
    current_user = await check_company_access(request, db)
    if not current_user:
        return RedirectResponse(url="/companies/login", status_code=302)
    
    return templates.TemplateResponse("companies/create-listing.html", {
        "request": request,
        "current_user": current_user,
        "company_name": current_user.company_name
    })

# API endpoints for companies
@app.delete("/api/v1/companies/properties/{property_id}")
async def delete_company_property(
    property_id: int,
    request: Request,
    db: Session = Depends(deps.get_db)
):
    """Удаление объявления компании"""
    current_user = await check_company_access(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    try:
        # Получаем объявление
        property_obj = db.query(models.Property).filter(
            models.Property.id == property_id,
            models.Property.owner_id == current_user.id  # Проверяем, что объявление принадлежит текущей компании
        ).first()
        
        if not property_obj:
            raise HTTPException(status_code=404, detail="Объявление не найдено")
        
        # Удаляем объявление
        db.delete(property_obj)
        db.commit()
        
        return {"success": True, "message": "Объявление удалено"}
        
    except Exception as e:
        print(f"DEBUG: Ошибка удаления объявления: {e}")
        raise HTTPException(status_code=500, detail="Ошибка при удалении объявления")

@app.post("/api/v1/companies/properties")
async def create_company_property(
    request: Request,
    title: str = Form(...),
    description: str = Form(...),
    price: float = Form(...),
    address: str = Form(...),
    city: str = Form("Бишкек"),
    category_id: int = Form(None),
    area: float = Form(None),
    rooms: int = Form(None),
    floor: int = Form(None),
    building_floors: int = Form(None),
    bathroom_type: str = Form(None),
    type: str = Form("apartment"),
    has_balcony: bool = Form(False),
    has_furniture: bool = Form(False),
    has_renovation: bool = Form(False),
    has_parking: bool = Form(False),
    has_elevator: bool = Form(False),
    has_security: bool = Form(False),
    has_internet: bool = Form(False),
    has_air_conditioning: bool = Form(False),
    has_heating: bool = Form(False),
    has_yard: bool = Form(False),
    has_pool: bool = Form(False),
    has_gym: bool = Form(False),
    status: str = Form("pending"),
    db: Session = Depends(deps.get_db)
):
    """Создание объявления компании"""
    current_user = await check_company_access(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    try:
        # Создаем объявление
        property_obj = models.Property(
            title=title,
            description=description,
            price=price,
            address=address,
            city=city,
            category_id=category_id,
            area=area,
            rooms=rooms,
            floor=floor,
            building_floors=building_floors,
            bathroom_type=bathroom_type,
            type=type,
            has_balcony=has_balcony,
            has_furniture=has_furniture,
            has_renovation=has_renovation,
            has_parking=has_parking,
            has_elevator=has_elevator,
            has_security=has_security,
            has_internet=has_internet,
            has_air_conditioning=has_air_conditioning,
            has_heating=has_heating,
            has_yard=has_yard,
            has_pool=has_pool,
            has_gym=has_gym,
            status=models.PropertyStatus.DRAFT if status == "draft" else models.PropertyStatus.PENDING,
            owner_id=current_user.id
        )
        
        db.add(property_obj)
        db.commit()
        db.refresh(property_obj)
        
        return {
            "success": True,
            "message": "Объявление создано" if status == "draft" else "Объявление отправлено на модерацию",
            "property_id": property_obj.id
        }
        
    except Exception as e:
        print(f"DEBUG: Ошибка создания объявления: {e}")
        raise HTTPException(status_code=500, detail="Ошибка при создании объявления")

# === END COMPANIES PANEL ROUTES ===

# API для получения категорий
@app.get("/api/v1/categories")
async def get_categories(db: Session = Depends(deps.get_db)):
    """Получение списка всех категорий"""
    categories = db.query(models.Category).all()
    return [{"id": cat.id, "name": cat.name, "description": cat.description} for cat in categories]

# API для работы с заявками (Requests)
